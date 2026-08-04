"""
logic.py — Shared game logic: valuation, evidence economy, dashboard aggregation,
and the editable schedule (round -> topic -> advance time).

Kept free of Streamlit so it can be unit-tested and reused.
"""

import json
import re
from datetime import datetime, timezone

import db
import content


# --------------------------------------------------------------------------- #
# Schedule: number of rounds, topic per round, advance datetimes, auto-advance
# --------------------------------------------------------------------------- #
def total_rounds():
    try:
        return int(db.get_setting("total_rounds", content.DEFAULT_TOTAL_ROUNDS))
    except (TypeError, ValueError):
        return content.DEFAULT_TOTAL_ROUNDS


def get_schedule():
    """Return the ordered schedule: each round with its LIST of topic dicts.

    A round may cover several pieces of material, which lets a full 15-topic
    curriculum be compressed into fewer rounds.
    """
    placements = db.get_round_topics()          # ordered by round, position
    adv = {r["round"]: r["advance_at"] for r in db.get_schedule_rows()}
    by_round = {}
    for p in placements:
        by_round.setdefault(p["round"], []).append(p["topic_key"])
    out = []
    n = total_rounds()
    for rnd in range(1, n + 1):
        keys = by_round.get(rnd, [])
        topics = [content.CURRICULUM_BY_KEY[k] for k in keys if k in content.CURRICULUM_BY_KEY]
        out.append({"round": rnd, "topics": topics, "advance_at": adv.get(rnd)})
    return out


def unassigned_topics():
    """Curriculum topics not placed in any round within range (the 'pool')."""
    n = total_rounds()
    placed = {p["topic_key"] for p in db.get_round_topics() if p["round"] and p["round"] <= n}
    return [t for t in content.CURRICULUM_TOPICS if t["key"] not in placed]


# --------------------------------------------------------------------------- #
# Balanced auto-arrangement
#   Each topic has a "load" (how much material it is). We split the ordered
#   curriculum into as many contiguous groups as there are rounds, minimizing the
#   heaviest round — so related concepts stay together and no round is overloaded.
# --------------------------------------------------------------------------- #
def topic_load(topic):
    """A rough effort weight for a topic: its concepts plus objectives."""
    if not topic:
        return 0
    return len(topic.get("concepts", [])) + len(topic.get("objectives", []))


def _balanced_group_sizes(weights, k):
    """Split a sequence of weights into k contiguous groups minimizing the max
    group sum (classic linear-partition DP). Returns a list of k group sizes.
    """
    n = len(weights)
    if k <= 1:
        return [n]
    if k >= n:
        return [1] * n + [0] * (k - n)
    prefix = [0]
    for w in weights:
        prefix.append(prefix[-1] + w)
    INF = float("inf")
    dp = [[INF] * (k + 1) for _ in range(n + 1)]
    back = [[0] * (k + 1) for _ in range(n + 1)]
    dp[0][0] = 0
    for i in range(1, n + 1):
        for j in range(1, min(i, k) + 1):
            for p in range(j - 1, i):
                seg = prefix[i] - prefix[p]
                cost = max(dp[p][j - 1], seg)
                if cost < dp[i][j]:
                    dp[i][j] = cost
                    back[i][j] = p
    sizes, i, j = [], n, k
    while j > 0:
        p = back[i][j]
        sizes.append(i - p)
        i, j = p, j - 1
    sizes.reverse()
    return sizes


def suggest_balanced_layout(n_rounds=None, order=None):
    """Suggest a balanced arrangement: {round: [topic_key, ...]}.

    Topics are taken in their logical curriculum order and split into contiguous
    groups so that related concepts stay together and the heaviest round is as
    light as possible.
    """
    n_rounds = int(n_rounds or total_rounds())
    order = order or content.DEFAULT_TOPIC_ORDER
    weights = [topic_load(content.CURRICULUM_BY_KEY[k]) for k in order]
    sizes = _balanced_group_sizes(weights, n_rounds)
    layout, idx = {}, 0
    for r, size in enumerate(sizes, start=1):
        layout[r] = order[idx:idx + size]
        idx += size
    return layout


def layout_load_summary(layout):
    """For a {round:[keys]} layout, return [{round, titles, load, count}]."""
    out = []
    for r in sorted(layout):
        keys = layout[r]
        out.append({
            "round": r,
            "titles": " + ".join(content.CURRICULUM_BY_KEY[k]["title"] for k in keys) or "—",
            "load": sum(topic_load(content.CURRICULUM_BY_KEY[k]) for k in keys),
            "count": len(keys),
        })
    return out


def apply_layout(layout):
    """Replace the whole schedule assignment with the given {round:[keys]} layout."""
    for tp in content.CURRICULUM_TOPICS:
        db.remove_round_topic(tp["key"])
    for r in sorted(layout):
        for pos, key in enumerate(layout[r]):
            db.set_topic_placement(key, r, pos)


def round_load(rnd):
    """Total load currently assigned to a round."""
    return sum(topic_load(t) for t in topics_for_round(rnd))


def set_total_rounds(n):
    """Change how many rounds the simulation runs.

    Growing adds empty rounds. Shrinking moves any material from removed rounds
    onto the last remaining round, so no topic is ever lost. Never lets
    current_round exceed the new max.
    """
    n = max(1, int(n))
    # Move orphaned material (round > n) onto the last round.
    for p in db.get_round_topics():
        if p["round"] and p["round"] > n:
            db.set_topic_placement(p["topic_key"], n)
    # Ensure advance-time rows exist for each round; drop extras.
    existing = {r["round"] for r in db.get_schedule_rows()}
    for rnd in range(1, n + 1):
        if rnd not in existing:
            db.upsert_schedule_row(rnd, None, None)
    db.delete_schedule_rows_above(n)
    db.set_setting("total_rounds", n)
    # Clamp every game's round to the new maximum.
    for g in db.list_games():
        if int(g["current_round"]) > n:
            db.set_game_round(g["id"], n)


def topics_for_round(rnd):
    """The list of curriculum topics covered in a given round."""
    for row in get_schedule():
        if row["round"] == rnd:
            return row["topics"]
    return []


def newly_unlocked(rnd):
    """Student tools first introduced at this round, across all its topics."""
    intro = []
    for topic in topics_for_round(rnd):
        for p in topic.get("introduces", []):
            if p not in intro:
                intro.append(p)
    if rnd == 1:
        intro = content.BASE_TOOLS + [p for p in intro if p not in content.BASE_TOOLS]
    return intro


def page_unlock_round(page):
    """Earliest round that introduces a page (1 for base tools / unscheduled)."""
    if page in content.BASE_TOOLS:
        return 1
    for row in get_schedule():
        for topic in row["topics"]:
            if page in topic.get("introduces", []):
                return row["round"]
    return 1  # never explicitly scheduled => always available


def canvas_unlock_round(canvas_type):
    """Earliest round whose material focuses on a given canvas type."""
    for row in get_schedule():
        for topic in row["topics"]:
            cs = topic.get("canvases") or ([topic["canvas"]] if topic.get("canvas") else [])
            if canvas_type in cs:
                return row["round"]
    return 1


def canvas_focus_for_round(rnd):
    """List of canvas types in focus this round (a round may cover several)."""
    focuses = []
    for topic in topics_for_round(rnd):
        cs = topic.get("canvases") or ([topic["canvas"]] if topic.get("canvas") else [])
        for c in cs:
            if c and c not in focuses:
                focuses.append(c)
    return focuses


def _total_rounds():
    sched = get_schedule()
    return max((row["round"] for row in sched), default=1)


def active_rounds_for_page(page):
    """Every round in which a tool is the active task (ignores carried-over work)."""
    return [r for r in range(1, _total_rounds() + 1)
            if page in active_tools(r, team_id=None)]


def active_rounds_for_canvas(ctype):
    """Every round in which a specific canvas is in focus."""
    return [r for r in range(1, _total_rounds() + 1)
            if ctype in canvas_focus_for_round(r)]


def rounds_for_topic(topic_key):
    """Every round whose material includes a given curriculum topic key."""
    out = []
    for row in get_schedule():
        if any(tp.get("key") == topic_key for tp in row["topics"]):
            out.append(row["round"])
    return out


def rounds_phrase(rounds):
    """Human phrase for a list of round numbers: 'Round 3', 'Rounds 3 and 5', etc."""
    rounds = sorted(set(rounds))
    if not rounds:
        return "a later round"
    if len(rounds) == 1:
        return f"Round {rounds[0]}"
    if len(rounds) == 2:
        return f"Rounds {rounds[0]} and {rounds[1]}"
    return "Rounds " + ", ".join(str(r) for r in rounds[:-1]) + f", and {rounds[-1]}"


# --------------------------------------------------------------------------- #
# Round deliverables, completion, and tool gating
# --------------------------------------------------------------------------- #
def _deliverable_done(team_id, check, rnd):
    if check == "always":
        return True
    if check == "ack_founder_review":
        return db.has_ack(team_id, "founder_review")
    if check == "time_plan_set":
        return db.has_ack(team_id, "time_plan_set")
    if check == "ventures_ge_3":
        return len(db.get_ventures(team_id)) >= 3
    if check == "cp_ge_1":
        return len(db.list_canvases(team_id, "customer_profile")) >= 1
    if check == "cp_ge_2":
        return len(db.list_canvases(team_id, "customer_profile")) >= 2
    if check == "vpc_ge_1":
        return len(db.list_canvases(team_id, "vpc")) >= 1
    if check == "bmc_ge_1":
        return len(db.list_canvases(team_id, "bmc")) >= 1
    if check == "bmc_ge_2":
        return len(db.list_canvases(team_id, "bmc")) >= 2
    if check == "bmc_ge_3":
        return len(db.list_canvases(team_id, "bmc")) >= 3
    if check == "env_ge_1":
        return len(db.list_canvases(team_id, "environment")) >= 1
    if check == "evidence_ge_2":
        return len(db.list_evidence(team_id)) >= 2
    if check == "evidence_ge_4":
        return len(db.list_evidence(team_id)) >= 4
    if check == "vps_ge_3":
        return len(db.list_value_props(team_id)) >= 3
    if check == "vp_results_ge_1":
        return len(db.list_vp_results(team_id)) >= 1
    if check == "assumptions_ge_5":
        return len(db.list_assumptions(team_id)) >= 5
    if check == "experiments_ge_2":
        return len(db.list_experiments(team_id)) >= 2
    if check == "experiment_results_ge_1":
        return any(e["outcome"] in ("Supported", "Refuted")
                   for e in db.list_experiments(team_id))
    if check == "pricing_exp":
        return any(("price" in (e["card_type"] or "").lower()
                    or "preorder" in (e["card_type"] or "").lower())
                   for e in db.list_experiments(team_id))
    if check == "pivots_ge_1":
        return len(db.list_pivots(team_id)) >= 1
    if check == "reflection_this_round":
        return any(r["round"] == rnd for r in db.list_reflections(team_id))
    if check == "ai_log_this_round":
        return any(l["round"] == rnd for l in db.list_ai_logs(team_id))
    return False


def round_requirements(rnd):
    """Deliverables a team must complete to finish this round (de-duplicated)."""
    reqs, seen = [], set()
    for tp in topics_for_round(rnd):
        for d in content.TOPIC_DELIVERABLES.get(tp["key"], []):
            if d["label"] not in seen:
                seen.add(d["label"])
                reqs.append(d)
    u = content.UNIVERSAL_DELIVERABLE
    if u["label"] not in seen:
        reqs.append(u)
    return reqs


def round_progress(team_id, rnd):
    """This round's own deliverables (decisions), each with a done flag."""
    return [{**d, "done": _deliverable_done(team_id, d["check"], rnd), "kind": "decision"}
            for d in round_requirements(rnd)]


# ---- Concept coverage (every concept covered by a decision or a question) --- #
def round_concepts(rnd):
    seen = []
    for tp in topics_for_round(rnd):
        for c in tp["concepts"]:
            if c not in seen:
                seen.append(c)
    return seen


# ---- Written-answer quality + concept quiz --------------------------------- #
_ANSWER_EVIDENCE_WORDS = {
    "evidence", "customer", "customers", "interview", "interviews", "test", "tested", "data",
    "tried", "paid", "signed", "observed", "survey", "result", "results", "behavior",
    "behaviour", "willingness", "preorder", "pilot", "experiment", "experiments", "proof",
}


def _quality_feedback(checks):
    msg = []
    if not checks["complete"]:
        msg.append("Write a bit more — a full sentence or two.")
    if not checks["meaningful"]:
        msg.append("Say something specific — this reads like filler or repetition.")
    if not checks["uses_concepts"]:
        msg.append("Use the round's concepts/terms (e.g. evidence, assumption, value, segment).")
    if not checks["relevant"]:
        msg.append("Tie it to YOUR venture/territory or to the concept itself.")
    if not checks["evidence_based"]:
        msg.append("Ground it in evidence — what customers did or said, a test, or data.")
    return msg


def answer_quality(text, concept="", team_id=None):
    """Check an open answer is complete, meaningful, concept-using, relevant, and
    evidence-based. Returns per-criterion flags + an overall ok. Pass rule: it must
    be complete and meaningful, and meet at least two of the three substance checks."""
    text = (text or "").strip()
    words = re.findall(r"[a-z']+", text.lower())
    n = len(words)
    uniq = set(words)
    checks = {}
    checks["complete"] = n >= 8
    checks["meaningful"] = (len(uniq) >= 5 and (len(uniq) / n if n else 0) > 0.4
                            and not re.search(r"(.)\1{4,}", text.lower()))
    checks["uses_concepts"] = bool(uniq & content.sim_vocab())
    ctx = _context_terms(team_id) if team_id is not None else set()
    cwords = set(re.findall(r"[a-z']{4,}", concept.lower()))
    checks["relevant"] = bool(uniq & ctx) or bool(uniq & cwords)
    checks["evidence_based"] = bool(uniq & _ANSWER_EVIDENCE_WORDS)
    substance = (checks["uses_concepts"] + checks["relevant"] + checks["evidence_based"])
    # Pass = a real, on-topic sentence: complete + meaningful + at least one substance signal.
    ok = checks["complete"] and checks["meaningful"] and substance >= 1
    return {"ok": ok, "checks": checks, "passed": sum(checks.values()), "total": 5,
            "feedback": _quality_feedback(checks)}


def answer_grade(text, concept="", team_id=None):
    """Accurate, graded feedback for a written concept answer — distinguishing blank,
    not-meaningful, developing, acceptable, and strong, instead of a generic 'almost'."""
    q = answer_quality(text, concept, team_id)
    c = q["checks"]
    t = (text or "").strip()
    if not t:
        return {"level": "blank", "icon": "⬜", "ok": False, "quality": q,
                "headline": "Blank — nothing written yet. Add a sentence or two applying "
                            "the idea to your venture."}
    if not c["complete"]:
        return {"level": "incomplete", "icon": "⬜", "ok": False, "quality": q,
                "headline": "Incomplete — too short. Give a full sentence or two."}
    if not c["meaningful"]:
        return {"level": "not_meaningful", "icon": "⚠️", "ok": False, "quality": q,
                "headline": "This isn't meaningful yet — it reads like filler or repeated "
                            "words. Say something specific about your venture."}
    substance = c["uses_concepts"] + c["relevant"] + c["evidence_based"]
    if substance == 0:
        return {"level": "developing", "icon": "🟡", "ok": False, "quality": q,
                "headline": "Getting there — now connect it to the course concepts, YOUR "
                            "venture, or your evidence."}
    missing = [lbl for k, lbl in (("uses_concepts", "the course concepts"),
                                  ("relevant", "your own venture"),
                                  ("evidence_based", "evidence")) if not c[k]]
    if not missing:
        return {"level": "strong", "icon": "🌟", "ok": True, "quality": q,
                "headline": "Excellent — complete, specific, and evidence-based."}
    if len(missing) == 1:
        join = missing[0]
    elif len(missing) == 2:
        join = " and ".join(missing)
    else:
        join = ", ".join(missing[:-1]) + ", and " + missing[-1]
    return {"level": "acceptable", "icon": "✅", "ok": True, "quality": q,
            "headline": "Acceptable — this counts. You could strengthen it by referencing "
                        + join + "."}


def _parse_concept_response(raw):
    """A stored concept answer may be JSON {quiz:[...], text:...} or legacy plain text."""
    raw = (raw or "").strip()
    if not raw:
        return {"quiz": None, "text": ""}
    if raw.startswith("{"):
        try:
            d = json.loads(raw)
            return {"quiz": d.get("quiz"), "text": d.get("text", "")}
        except (ValueError, TypeError):
            pass
    return {"quiz": None, "text": raw}


def concept_quiz_correct(concept, quiz_answers):
    """Whether the team's true/false picks match the concept's quiz (True if no quiz)."""
    qs = content.CONCEPT_QUIZ.get(concept)
    if not qs:
        return True
    if not quiz_answers or len(quiz_answers) != len(qs):
        return False
    return all(bool(a) == bool(truth) for (_, truth), a in zip(qs, quiz_answers))


def concept_answer_status(team_id, rnd, concept, answers=None):
    """Full status of a written concept: parsed response, quiz correctness, answer
    quality, and whether it's DONE (quiz correct AND answer passes quality)."""
    if answers is None:
        answers = db.get_round_answers(team_id, rnd)
    r = _parse_concept_response(answers.get(concept))
    quality = answer_quality(r["text"], concept, team_id)
    quiz_ok = concept_quiz_correct(concept, r["quiz"])
    return {"text": r["text"], "quiz": r["quiz"], "quiz_ok": quiz_ok,
            "quality": quality, "done": quiz_ok and quality["ok"]}


def concept_progress(team_id, rnd):
    """Each concept this round with how it's covered.

    A concept that maps to a decision (CONCEPT_CHECKS) is covered the moment that
    decision is done — no written answer required. Concepts with no such decision
    are checked with a short true/false quiz plus a quality-checked applied answer."""
    answers = db.get_round_answers(team_id, rnd)
    reqs = {d["check"]: d for d in round_requirements(rnd)}
    out = []
    for c in round_concepts(rnd):
        chk = content.CONCEPT_CHECKS.get(c)
        if chk:
            d = reqs.get(chk, {})
            out.append({"concept": c, "label": c, "kind": "decision", "check": chk,
                        "tool": d.get("tool", ""), "action": d.get("label", ""),
                        "needs_question": False, "must_update": True,
                        "done": _deliverable_done(team_id, chk, rnd)})
        else:
            stt = concept_answer_status(team_id, rnd, c, answers)
            out.append({"concept": c, "label": f"Concept check — {c}",
                        "tool": "Concept Check", "kind": "question",
                        "needs_question": True, "must_update": True,
                        "has_quiz": bool(content.CONCEPT_QUIZ.get(c)),
                        "done": stt["done"]})
    return out


# ---- Carry-forward of unfinished prior work -------------------------------- #
def outstanding_prior(team_id, rnd):
    """Incomplete deliverables AND unanswered concept-checks from earlier rounds."""
    out = []
    for r in range(1, rnd):
        for d in round_requirements(r):
            if not _deliverable_done(team_id, d["check"], r):
                out.append({**d, "round": r, "carried": True, "done": False,
                            "kind": "decision"})
        answers = db.get_round_answers(team_id, r)
        for c in round_concepts(r):
            # Concepts covered by a decision are carried via that decision above —
            # only genuine open-ended questions are carried here.
            if content.CONCEPT_CHECKS.get(c):
                continue
            if not concept_answer_status(team_id, r, c, answers)["done"]:
                out.append({"concept": c, "label": f"Concept check — {c}",
                            "tool": "Concept Check", "round": r, "carried": True,
                            "kind": "question", "must_update": True, "done": False})
    return out


def round_checklist(team_id, rnd):
    """Everything a team must finish THIS round: this round's decisions and concept
    questions, plus any unfinished work carried over from earlier rounds."""
    return {
        "decisions": round_progress(team_id, rnd),
        "questions": concept_progress(team_id, rnd),
        "carried": outstanding_prior(team_id, rnd),
    }


def round_complete(team_id, rnd):
    cl = round_checklist(team_id, rnd)
    items = cl["decisions"] + cl["questions"] + cl["carried"]
    return all(i["done"] for i in items) if items else True


def round_progress_counts(team_id, rnd):
    """(done, total) across everything the team must finish this round."""
    cl = round_checklist(team_id, rnd)
    items = cl["decisions"] + cl["questions"] + cl["carried"]
    return sum(1 for i in items if i["done"]), len(items)


def next_action(team_id, rnd):
    """The single most important thing to do next this round, or None if all done.

    Priority: clear carried-over backlog first, then this round's decisions, then any
    concept questions — so a team always has one clear next move."""
    cl = round_checklist(team_id, rnd)
    for bucket in (cl["carried"], cl["decisions"], cl["questions"]):
        for it in bucket:
            if not it["done"]:
                return {"label": it.get("label", it.get("concept", "")),
                        "tool": it.get("tool", ""), "carried": bool(it.get("carried"))}
    return None


def strict_round_mode():
    """When on (default), tools not relevant to the current round are view-only."""
    return auto_flag("strict_round_mode", default=True)


def active_tools(rnd, team_id=None):
    """Student tools that are editable/relevant this round (incl. carried-over work)."""
    tools = set(content.ALWAYS_ACTIVE_TOOLS)
    for tp in topics_for_round(rnd):
        if tp.get("tool"):
            tools.add(tp["tool"])
        for p in tp.get("introduces", []):
            tools.add(p)
        if tp.get("canvas"):
            tools.add("Canvases")
        for d in content.TOPIC_DELIVERABLES.get(tp["key"], []):
            if d.get("tool"):
                tools.add(d["tool"])
    if team_id is not None:
        for d in outstanding_prior(team_id, rnd):
            if d.get("tool"):
                tools.add(d["tool"])
    return tools


def tool_state(page, rnd, team_id=None):
    """'locked' (not introduced), 'active' (relevant now / carried work), or 'reference'."""
    if page_unlock_round(page) > rnd:
        return "locked"
    if page in active_tools(rnd, team_id):
        return "active"
    return "reference"


def tool_editable(page, rnd, team_id=None):
    state = tool_state(page, rnd, team_id)
    if state == "locked":
        return False
    if state == "reference":
        return not strict_round_mode()
    return True


# ---- Per-canvas gating: a canvas is editable only in its own round --------- #
_CHECK_CANVAS = {
    "cp_ge_1": "customer_profile", "cp_ge_2": "customer_profile",
    "vpc_ge_1": "vpc",
    "bmc_ge_1": "bmc", "bmc_ge_2": "bmc", "bmc_ge_3": "bmc",
    "env_ge_1": "environment",
}


def editable_canvas_types(team_id, rnd):
    """Canvas types the team may edit this round: this round's focus canvases plus
    any canvas whose deliverable is still carried over from an earlier round."""
    types = set(canvas_focus_for_round(rnd))
    if team_id is not None:
        for d in outstanding_prior(team_id, rnd):
            c = _CHECK_CANVAS.get(d.get("check"))
            if c:
                types.add(c)
    return types


def canvas_unlock_round_for(ctype):
    return canvas_unlock_round(ctype)


def canvas_editable(ctype, rnd, team_id=None):
    """Whether a specific canvas type can be edited this round."""
    if not tool_editable("Canvases", rnd, team_id):
        return False
    if not strict_round_mode():
        return True
    return ctype in editable_canvas_types(team_id, rnd)


# ---- Founder / team skills ------------------------------------------------- #
def hire_boost(team_id):
    """Total hired boost per skill key."""
    out = {}
    for h in db.list_hires(team_id):
        out[h["skill_key"]] = out.get(h["skill_key"], 0) + (h["boost"] or 0)
    return out


def effective_skill(team_id, skill_key, skills=None, boosts=None):
    """Founder level + hired boost for a skill, capped at HIRE_SKILL_CAP."""
    skills = skills if skills is not None else db.get_team_skills(team_id)
    boosts = boosts if boosts is not None else hire_boost(team_id)
    return min(content.HIRE_SKILL_CAP, skills.get(skill_key, 0) + boosts.get(skill_key, 0))


def skill_bonus(team_id):
    """Score bonus per dashboard dimension from EFFECTIVE skills (founder + hires)."""
    bonus = {}
    skills = db.get_team_skills(team_id)
    boosts = hire_boost(team_id)
    for s in content.FOUNDER_SKILLS:
        lvl = effective_skill(team_id, s["key"], skills, boosts)
        bonus[s["dimension"]] = bonus.get(s["dimension"], 0) + lvl * 3
    return bonus


# ---- Economy / balance settings (Director-tunable, with defaults) ---------- #
_ECON_DEFAULTS = {
    "train_mult": 10.0,
    "pt_boost": 2, "pt_upfront": 150.0, "pt_per_round": 80.0,
    "ft_boost": 3, "ft_upfront": 400.0, "ft_per_round": 200.0,
}


def _econ(key):
    raw = db.get_setting("econ_" + key)
    if raw in (None, ""):
        return _ECON_DEFAULTS[key]
    try:
        return type(_ECON_DEFAULTS[key])(float(raw))
    except (TypeError, ValueError):
        return _ECON_DEFAULTS[key]


def get_economy():
    """Current economy config (Director overrides merged over defaults)."""
    return {k: _econ(k) for k in _ECON_DEFAULTS}


def set_economy(values):
    for k in _ECON_DEFAULTS:
        if k in values and values[k] is not None:
            db.set_setting("econ_" + k, values[k])


def hire_options():
    """Hire options built from the current economy settings."""
    e = get_economy()
    return {
        "part_time": {"label": "Part-time", "boost": int(e["pt_boost"]),
                      "upfront": e["pt_upfront"], "per_round": e["pt_per_round"]},
        "full_time": {"label": "Full-time", "boost": int(e["ft_boost"]),
                      "upfront": e["ft_upfront"], "per_round": e["ft_per_round"]},
    }


def skill_train_cost(level):
    """Founder-hours to raise a skill from `level` to level+1 (rising cost)."""
    return int((level + 1) * _econ("train_mult"))


def cost_to_next(level):
    """Effective hours to raise a skill from `level` to level+1."""
    return skill_train_cost(level)


def skill_progress(team_id, skill_key):
    """(current_level, banked_xp_hours, cost_to_next_level)."""
    lvl = db.get_team_skills(team_id).get(skill_key, 0)
    xp = db.get_skill_xp(team_id, skill_key)
    return lvl, xp, cost_to_next(lvl)


def add_skill_progress(team_id, skill_key, eff_hours):
    """Bank effective-hours of progress toward a skill; level up while it covers the
    cost. Partial progress is kept (never wasted). Returns levels gained."""
    xp = db.get_skill_xp(team_id, skill_key) + eff_hours
    lvl = db.get_team_skills(team_id).get(skill_key, 0)
    gained = 0
    while lvl < content.SKILL_MAX and xp >= cost_to_next(lvl):
        xp -= cost_to_next(lvl)
        lvl += 1
        gained += 1
    if lvl >= content.SKILL_MAX:
        xp = 0
    db.set_skill_level(team_id, skill_key, lvl)
    db.set_skill_xp(team_id, skill_key, xp)
    return gained


def invest_training(team_id, skill_key, hours):
    """Assign founder time to training a skill. The hours count toward this round's
    effort (capped at 80) and bank exactly as XP toward the next level — partial
    effort carries over to future rounds. Returns (ok, message)."""
    hours = int(hours)
    if hours <= 0:
        return False, "Enter a positive number of hours."
    lvl = db.get_team_skills(team_id).get(skill_key, 0)
    if lvl >= content.SKILL_MAX:
        return False, "Already at maximum level."
    room = effort_headroom(db.get_team(team_id))
    if hours > room:
        return False, (f"That would push the founder over 80 hours. Only {room}h of effort "
                       "left this week — free some up or work fewer tasks.")
    _add_usage(team_id, "spent_train", hours)
    gained = add_skill_progress(team_id, skill_key, hours)
    name = content.FOUNDER_SKILL_BY_KEY[skill_key]["name"]
    _, xp, nxt = skill_progress(team_id, skill_key)
    if gained:
        return True, f"Assigned {hours}h — {name} rose {gained} level(s)! ({xp}/{nxt} toward next)."
    return True, f"Assigned {hours}h — banked toward {name} ({xp}/{nxt} to next level)."


def _card_base_level(team_id, skill_key):
    card = db.get_founder_card(team_id)
    return content.card_skill_levels(card.get("name", "")).get(skill_key, 1)


def untrain_skill(team_id, skill_key):
    """Change your mind: first drop any banked (not-yet-levelled) training progress
    for this skill, freeing that effort back up; otherwise revert one trained level
    (down to the founder card's starting level). Learning-by-doing and starting
    levels are kept."""
    lvl = db.get_team_skills(team_id).get(skill_key, 0)
    xp = db.get_skill_xp(team_id, skill_key)
    base = _card_base_level(team_id, skill_key)
    name = content.FOUNDER_SKILL_BY_KEY[skill_key]["name"]
    if xp > 0:
        # free the effort back (reduce this round's training usage)
        freed = min(int(db.get_team(team_id).get("spent_train") or 0), xp)
        if freed:
            _add_usage(team_id, "spent_train", -freed)
        db.set_skill_xp(team_id, skill_key, 0)
        return True, f"Dropped {xp}h of banked training for {name} (freed {freed}h of effort)."
    if lvl > base:
        db.set_skill_level(team_id, skill_key, lvl - 1)
        return True, f"{name} reverted to level {lvl - 1}."
    return False, "Nothing to undo — this is the founder's starting level."


# ---- Hiring ---------------------------------------------------------------- #
def hire_specialist(team_id, skill_key, kind):
    """Hire a specialist. Costs money (upfront) AND founder time (recruiting), and
    adds an ongoing salary (money) plus management time (hours) each round."""
    opt = hire_options().get(kind)
    if not opt:
        return False, "Unknown employment type."
    role = content.SPECIALIST_ROLES.get(skill_key, "Specialist")
    recruit = content.HIRE_OPTIONS[kind].get("recruit_hours", 0)
    manage = content.HIRE_OPTIONS[kind].get("manage_hours", 0)
    ok, msg = db.adjust_resources(
        team_id, money=-opt["upfront"], hours=-recruit, kind="hire",
        description=f"Hired {opt['label']} {role}")
    if not ok:
        return False, f"{msg} (hiring needs ${opt['upfront']:.0f} and {recruit}h to recruit.)"
    db.add_hire(team_id, skill_key, role, kind, opt["boost"], opt["per_round"], manage)
    if recruit:
        _add_usage(team_id, "spent_other", recruit)
    return True, (f"Hired a {opt['label'].lower()} {role} (+{opt['boost']} {role}). "
                  f"Ongoing: ${opt['per_round']:.0f}/round and {manage}h/round to manage.")


def fire_specialist(hire_id):
    db.remove_hire(hire_id)


def management_hours(team_id):
    """Founder hours per round consumed managing current hires."""
    return sum((h.get("manage_hours") or 0) for h in db.list_hires(team_id))


def skills_needed_this_round(rnd):
    """Founder skills the current round leans on most (union across its topics)."""
    needs = []
    for tp in topics_for_round(rnd):
        for k in content.TOPIC_SKILL_NEEDS.get(tp["key"], []):
            if k not in needs:
                needs.append(k)
    return needs


# ---- Hours model: weekly budget, reset each round, diminishing returns ------ #
def productive_hours(raw):
    """Effective founder-hours from a raw weekly figure. Hours up to 40 are fully
    productive; each hour beyond 40 (to a max of 80) counts for less."""
    raw = max(0, min(content.MAX_WEEKLY_HOURS, raw or 0))
    if raw <= content.FULL_PRODUCTIVITY_HOURS:
        return int(round(raw))
    over = raw - content.FULL_PRODUCTIVITY_HOURS
    return int(round(content.FULL_PRODUCTIVITY_HOURS + over * content.OVERWORK_PRODUCTIVITY))


# --------------------------------------------------------------------------- #
# Time model — the founder's weekly EFFORT is an OUTPUT of how the team assigns
# time to tasks, not a chosen slider. Effort = admin (grows with rounds) +
# managing hires + business-development budget + training + hiring time. It is
# hard-capped at 80 hours and colour-coded (green ≤40, yellow ≤60, red ≤80).
# --------------------------------------------------------------------------- #
def admin_hours(team):
    """Unavoidable admin/overhead time this round — grows as the venture (round)
    gets more complex."""
    rnd = db.current_round()
    return int(min(content.ADMIN_MAX_HOURS, content.ADMIN_BASE_HOURS + (rnd - 1) // 2))


def build_budget(team):
    """Hours the team has committed to business development (experiments) this round."""
    b = team.get("build_budget")
    return int(b) if b is not None else default_build(team)


def current_effort(team):
    """Total founder hours committed this round (the dynamic 'effort')."""
    return int(admin_hours(team) + management_hours(team["id"]) + build_budget(team)
               + (team.get("spent_train") or 0) + (team.get("spent_other") or 0))


def effort_headroom(team):
    """Founder hours still available before hitting the 80-hour cap."""
    return max(0, content.MAX_WEEKLY_HOURS - current_effort(team))


def max_build(team):
    """Largest business-dev budget that keeps total effort within 80 hours."""
    return max(0, content.MAX_WEEKLY_HOURS - admin_hours(team)
               - int(management_hours(team["id"])) - int(team.get("spent_train") or 0)
               - int(team.get("spent_other") or 0))


def default_build(team):
    """A sensible starting business-dev budget (keeps effort near the green line)."""
    fixed = admin_hours(team) + int(management_hours(team["id"]))
    return max(0, min(max_build(team), content.EFFORT_GREEN - fixed))


def effort_color(effort):
    """(css_color, emoji, label) for a total effort level."""
    if effort <= content.EFFORT_GREEN:
        return "#2b9d8f", "🟢", "sustainable"
    if effort <= content.EFFORT_YELLOW:
        return "#e0a000", "🟡", "stretched"
    return "#d9534f", "🔴", "overwork"


def set_build_budget(team_id, hours):
    """Commit business-development hours this round (capped so effort ≤ 80)."""
    team = db.get_team(team_id)
    hours = int(max(0, min(max_build(team), hours)))
    spent_b = int(team.get("spent_build") or 0)
    db.update_team(team_id, build_budget=hours,
                   founder_hours=max(0, hours - spent_b), hours_round=db.current_round())
    db.set_ack(team_id, "time_plan_set")


def sync_round_hours(team):
    """Reset the round: default the business-dev budget, refill building hours, and
    clear per-round usage counters (unused hours are lost). Runs once per round."""
    cur = db.current_round()
    if int(team.get("hours_round") or 0) != cur:
        b = default_build(team)
        db.update_team(team["id"], build_budget=b, founder_hours=b, hours_round=cur,
                       spent_build=0, spent_train=0, spent_other=0)
        return db.get_team(team["id"])
    return team


def _add_usage(team_id, field, hours):
    """Increment a per-round hours-usage counter (spent_build/train/other)."""
    cur = db.get_team(team_id).get(field) or 0
    db.update_team(team_id, **{field: cur + hours})


def hours_breakdown(team):
    """The founder's committed week by task (sums to current effort), plus each
    hire's own working hours. Returns [(label, hours)]."""
    cats = [
        ("Admin", admin_hours(team)),
        ("Managing hires", int(management_hours(team["id"]))),
        ("Business dev", build_budget(team)),
        ("Training", int(team.get("spent_train") or 0)),
        ("Hiring", int(team.get("spent_other") or 0)),
    ]
    for h in db.list_hires(team["id"]):
        kind = "FT" if h["kind"] == "full_time" else "PT"
        cats.append((f"{kind} {h['role']}",
                     int(content.HIRE_OPTIONS.get(h["kind"], {}).get("work_hours", 0))))
    return cats


# Legacy aliases kept so other modules/founder card keep working.
def weekly_hours(team):
    return current_effort(team)


def round_hours_budget(team):
    return build_budget(team)


# ---- Founder learning-by-doing --------------------------------------------- #
def round_own_complete(team_id, rnd):
    """Whether a team finished THAT round's own decisions and concept checks
    (ignoring backlog carried from earlier rounds)."""
    items = round_progress(team_id, rnd) + concept_progress(team_id, rnd)
    return all(i["done"] for i in items) if items else False


def _apply_round_learning(team_id, completed_round):
    """Founder learns by running the business: doing a round's own work grows the
    founder skills that round leaned on (learning tied to the team's progress)."""
    if not round_own_complete(team_id, completed_round):
        return
    for k in skills_needed_this_round(completed_round):
        add_skill_progress(team_id, k, content.LEARNING_HOURS_PER_ROUND)


# ---- Per-round reset, salary, and learning --------------------------------- #
def on_round_change(new_round):
    """When the round advances: apply learning for each completed round, RESET each
    team's founder-hours to this week's budget (no carryover — unused hours are
    lost), and charge specialist salaries. Idempotent via a per-game marker."""
    gid = db.active_game_id()
    if gid:
        marker = db.get_game_marker(gid)
    else:
        try:
            marker = int(db.get_setting("hours_marker", "1"))
        except (TypeError, ValueError):
            marker = 1
    if new_round <= marker:
        return
    rounds_passed = new_round - marker
    for t in db.list_teams():   # active game's teams (or all, if no game context)
        # snapshot each finished round's learning metrics BEFORE applying new learning,
        # so the team can see its progress trend over time.
        for r in range(marker, new_round):
            snapshot_metrics(t["id"], r)
        # learning from each round that just finished
        for r in range(marker, new_round):
            _apply_round_learning(t["id"], r)
        # (available hours reset lazily via sync_round_hours when a team next loads)
        # charge salaries for each round passed
        salary = sum((h["per_round"] or 0) for h in db.list_hires(t["id"]))
        if salary:
            db.adjust_resources(t["id"], money=-salary * rounds_passed, kind="salary",
                                description="Specialist salaries", allow_negative=True)
    if gid:
        db.set_game_marker(gid, new_round)
    else:
        db.set_setting("hours_marker", new_round)


def _parse_dt(text):
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def maybe_auto_advance():
    """Advance the current round if a scheduled advance time has passed.

    The app has no background scheduler, so this runs on page load: current_round
    becomes the highest round whose advance_at is in the past (never going backward,
    never exceeding total_rounds). Returns the (possibly unchanged) current round.
    """
    now = datetime.now(timezone.utc)
    cur = db.current_round()
    target = cur
    for row in get_schedule():
        dt = _parse_dt(row["advance_at"])
        if dt is not None:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if dt <= now and row["round"] > target:
                target = row["round"]
    if target != cur:
        db.set_current_round(target)
    return target


def next_scheduled_advance():
    """Return (round, datetime_text) for the next future advance, or None."""
    now = datetime.now(timezone.utc)
    for row in get_schedule():
        dt = _parse_dt(row["advance_at"])
        if dt is not None:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if dt > now:
                return row["round"], row["advance_at"]
    return None


# --------------------------------------------------------------------------- #
# Decision deadlines & round commitments
#   A round's decisions are due when the NEXT round begins (its scheduled advance
#   time). If that time isn't set, the deadline is "unspecified" and the round
#   never auto-locks — the Director decides when to move on.
# --------------------------------------------------------------------------- #
def _aware(dt):
    if dt is not None and dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def decision_due_at(rnd):
    """ISO text of when round `rnd`'s decisions are due, or None if unspecified.

    Uses the start time of the next round (round rnd+1's advance_at)."""
    for row in get_schedule():
        if row["round"] == rnd + 1:
            return row["advance_at"]
    return None


def decision_due_dt(rnd):
    return _aware(_parse_dt(decision_due_at(rnd)))


def deadline_status(rnd):
    """Human-friendly deadline info for a round.

    Returns a dict: {set: bool, due_at: iso|None, due_text: str, passed: bool,
    remaining: str}."""
    due = decision_due_dt(rnd)
    if due is None:
        return {"set": False, "due_at": None,
                "due_text": "No deadline set — decisions stay open until your instructor "
                            "advances the round.",
                "passed": False, "remaining": ""}
    now = datetime.now(timezone.utc)
    passed = due <= now
    if passed:
        remaining = "past due"
    else:
        delta = due - now
        days = delta.days
        hours = delta.seconds // 3600
        mins = (delta.seconds % 3600) // 60
        remaining = (f"{days}d {hours}h left" if days else
                     (f"{hours}h {mins}m left" if hours else f"{mins}m left"))
    return {"set": True, "due_at": decision_due_at(rnd),
            "due_text": due.astimezone().strftime("%a %b %d, %Y at %I:%M %p"),
            "passed": passed, "remaining": remaining}


def commitment_snapshot(team_id, rnd):
    """A compact record of what is / isn't done for this round, for scoring."""
    cl = round_checklist(team_id, rnd)
    items = cl["decisions"] + cl["questions"] + cl["carried"]
    done = sum(1 for i in items if i["done"])
    return {"total": len(items), "done": done,
            "labels_done": [i.get("label") for i in items if i["done"]],
            "labels_open": [i.get("label") for i in items if not i["done"]]}


def commitment_state(team_id, rnd):
    """Full commit state for a team+round: committed?, locked (past due)?, snapshot."""
    row = db.get_commitment(team_id, rnd)
    ds = deadline_status(rnd)
    committed = bool(row and row["committed"])
    # Once the deadline passes, a team can no longer change its commitment.
    locked = ds["set"] and ds["passed"]
    return {"committed": committed, "locked": locked,
            "committed_at": (row or {}).get("committed_at"),
            "deadline": ds}


def editing_locked(team_id, rnd=None):
    """True when the team's work is frozen for the round — i.e. they've committed.

    While committed, every tool is view-only so a team can't quietly change committed
    work and forget to re-commit. Withdrawing (before the deadline) unlocks editing."""
    rnd = db.current_round() if rnd is None else rnd
    return commitment_state(team_id, rnd)["committed"]


def commit_round(team_id, rnd):
    """Team locks in this round's work. No-op if the deadline has passed already."""
    st = commitment_state(team_id, rnd)
    if st["locked"]:
        return False, "The deadline has passed — this round is locked."
    snap = commitment_snapshot(team_id, rnd)
    db.set_commitment(team_id, rnd, True, decision_due_at(rnd), json.dumps(snap))
    return True, "Committed."


def decommit_round(team_id, rnd):
    """Team withdraws its commitment to keep editing — only before the deadline."""
    st = commitment_state(team_id, rnd)
    if st["locked"]:
        return False, "The deadline has passed — you can no longer withdraw."
    db.set_commitment(team_id, rnd, False, decision_due_at(rnd), None)
    return True, "Commitment withdrawn — you can keep editing."


# --------------------------------------------------------------------------- #
# Evidence economy
# --------------------------------------------------------------------------- #
def credits_for_evidence(strength):
    """Evidence Credits earned for a logged piece of evidence of a given strength."""
    return round(strength * content.CREDITS_PER_STRENGTH, 1)


# Language cues used to sanity-check how a team rated a piece of evidence.
# Behavior beats opinion: the ladder pays for what customers DID, not what they SAID.
_OPINION_CUES = [
    "would", "could", "might", "may ", "think", "thinks", "thought", "believe",
    "like the idea", "love the idea", "likes it", "loved it", "interested",
    "sounds ", "seems ", "probably", "maybe", "i'd", "they'd", "we'd",
    "would pay", "would buy", "would use", "would try", "plan to", "planning to",
    "intend", "intention", "hypothetical", "in theory", "cool", "nice idea",
    "great idea", "said they", "told us they", "wants ", "want to", "hope",
]
_BEHAVIOR_CUES = [
    "paid", "pay us", "bought", "purchase", "purchased", "signed", "sign-up",
    "signed up", "preorder", "pre-order", "pre-ordered", "preordered", "deposit",
    "subscribed", "subscription", "returned", "reordered", "used it", "trial",
    "trialed", "downloaded", "installed", "committed", "letter of intent", "loi",
    "invoice", "wired", "venmo", "credit card", "put down", "gave us money",
]


def _has_cue(text, cues):
    t = (text or "").lower()
    return any(c in t for c in cues)


def evidence_flags(description, source, strength):
    """Heuristic misclassification check comparing the WORDS to the claimed strength.

    Returns a list of gentle warning strings (empty if nothing looks off). The goal
    is to make teams practice telling behavior from opinion — the core skill — not to
    block them."""
    flags = []
    text = f"{description or ''} {source or ''}"
    opinion = _has_cue(text, _OPINION_CUES)
    behavior = _has_cue(text, _BEHAVIOR_CUES)
    # Claimed as behavioral/committal but the wording reads like a stated intention.
    if strength >= 6 and opinion and not behavior:
        flags.append(
            "This reads like something a customer **said** or **would** do, not something "
            f"they actually did — behavioral evidence (strength {strength}) usually describes "
            "an action taken (paid, signed, pre-ordered, used). Double-check the rating, or "
            "add what they physically did.")
    # A hypothetical/intention rating is fine, but nudge toward stronger tests.
    if strength <= 2 and behavior and not opinion:
        flags.append(
            "This describes a real action, which is usually **stronger** than a strength-"
            f"{strength} opinion — you may be under-crediting it. Re-check the ladder.")
    # High strength claimed with no concrete action words at all.
    if strength >= 8 and not behavior:
        flags.append(
            f"Strength {strength} is near the top of the ladder (a binding commitment). Make "
            "sure the description names the concrete commitment (a signed LOI, a payment, a "
            "paid trial) — reviewers and your instructor will look for it.")
    return flags


def log_evidence_and_award(team_id, description, evidence_type, source,
                           assumption_id=None, justification=None):
    """Add evidence to the ledger and pay Evidence Credits for it.

    Behavioral evidence is worth more, so credits scale with ladder strength.
    Returns (credits_awarded, strength).
    """
    strength = content.EVIDENCE_LADDER_MAP.get(evidence_type, 0)
    award = credits_for_evidence(strength)
    db.add_evidence(team_id, description, evidence_type, strength, source,
                    assumption_id, award, justification)
    if award > 0:
        db.adjust_resources(
            team_id, credits=award, kind="evidence",
            description=f"Evidence logged: {evidence_type}", allow_negative=True,
        )
    return award, strength


# --------------------------------------------------------------------------- #
# Experiment purchasing
# --------------------------------------------------------------------------- #
def purchase_experiment(team_id, card, assumption_id, hypothesis, metric,
                        success_threshold, failure_threshold, decision_rule,
                        predicted_outcome=None, confidence=None):
    """Charge the team and record a designed experiment.

    Captures the team's PREDICTION and confidence up front so we can later compare
    forecast vs. result — a calibration loop that builds entrepreneurial judgment.
    Returns (ok, message, experiment_id|None).
    """
    ok, msg = db.adjust_resources(
        team_id,
        money=-card["money"],
        credits=-card["credits"],
        hours=-card["hours"],
        kind="experiment",
        description=f"Purchased experiment: {card['name']}",
    )
    if not ok:
        return False, msg, None
    exp_id = db.add_experiment(
        team_id, assumption_id, card["name"], card["money"], card["hours"],
        card["credits"], card["strength"], hypothesis, metric,
        success_threshold, failure_threshold, decision_rule,
        predicted_outcome=predicted_outcome, confidence=confidence,
    )
    _add_usage(team_id, "spent_build", card["hours"])
    return True, "Experiment purchased and designed.", exp_id


# --------------------------------------------------------------------------- #
# Valuation
#   Potential value = Market Potential x Evidence Confidence x BM Coherence
#                     x Execution Factor              (what it COULD be worth)
#   Venture value   = Potential value x Evidence Coverage - Unresolved Risk
#   Each index runs 0.50–1.50 and is derived from dashboard scores (0–100).
#   Evidence coverage (0–1) discounts the opportunity by how much of the model
#   the team has actually BACKED WITH EVIDENCE — so an unproven idea is worth
#   little in week 1 and its value is EARNED as evidence comes in.
# --------------------------------------------------------------------------- #
def _index_from_score(score, default=0.90):
    """Map a 0–100 dashboard score onto a 0.50–1.50 index."""
    if score is None:
        return default
    return round(0.50 + (max(0.0, min(100.0, score)) / 100.0), 3)


# Evidence-coverage tuning.
COVERAGE_FLOOR = 0.0           # an unproven venture is worth NOTHING until work is done
COVERAGE_STRENGTH_TARGET = 40  # total evidence strength that counts as "well evidenced"
COVERAGE_IMPORTANCE_MIN = 3    # assumptions this important+ count toward coverage
COVERAGE_BUILD_TARGET = 8      # canvas versions + assumptions that count as "model built"


def evidence_coverage(team_id):
    """0–1 factor: how much of the venture is actually WORTH so far.

    Starts at exactly 0 for a brand-new team (a good idea with no work is worth
    nothing) and rises with the decisions they make and the evidence they gather:
      • model-building work — drafting canvases and naming assumptions (small)
      • tested important assumptions — reducing real uncertainty (large)
      • the strength of the evidence portfolio — behaviour beats opinion (large)
    """
    assums = db.list_assumptions(team_id)
    important = [a for a in assums if a["importance"] >= COVERAGE_IMPORTANCE_MIN]
    if important:
        tested = sum(1 for a in important if a["status"] in ("Supported", "Refuted"))
        tested_ratio = tested / len(important)
    else:
        tested_ratio = 0.0
    ev = evidence_summary(team_id)
    strength_factor = min(1.0, ev["total_strength"] / COVERAGE_STRENGTH_TARGET)
    # Model-development signal: building canvases and identifying assumptions is real
    # work that earns a little value even before anything is tested.
    build_units = len(db.list_canvases(team_id)) + len(assums)
    build_factor = min(1.0, build_units / COVERAGE_BUILD_TARGET)
    coverage = 0.5 * tested_ratio + 0.35 * strength_factor + 0.15 * build_factor
    return round(max(COVERAGE_FLOOR, min(1.0, coverage)), 3)


def compute_valuation(team_id):
    """Return a dict with the valuation and its components."""
    team = db.get_team(team_id)
    if not team:
        return None
    scores = db.latest_scores(team_id)

    market_potential = team["market_potential"] or 0
    evidence_conf = _index_from_score(scores.get("Evidence Strength"))
    bm_coherence = _index_from_score(scores.get("Business-Model Coherence"))
    execution = _index_from_score(scores.get("Team Execution"))
    unresolved_risk = team["unresolved_risk"] or 0

    potential = market_potential * evidence_conf * bm_coherence * execution
    coverage = evidence_coverage(team_id)
    value = potential * coverage - unresolved_risk
    return {
        "market_potential": market_potential,
        "evidence_confidence": evidence_conf,
        "bm_coherence": bm_coherence,
        "execution_factor": execution,
        "unresolved_risk": unresolved_risk,
        "evidence_coverage": coverage,
        "potential_valuation": round(potential, 0),
        "valuation": round(value, 0),
    }


# --------------------------------------------------------------------------- #
# Evidence-portfolio analytics
# --------------------------------------------------------------------------- #
def evidence_summary(team_id):
    ev = db.list_evidence(team_id)
    total_strength = sum(e["strength"] for e in ev)
    total_credits = sum(e["credits_award"] for e in ev)
    behavioral = sum(1 for e in ev if e["strength"] >= 6)
    opinion = sum(1 for e in ev if e["strength"] <= 2)
    return {
        "count": len(ev),
        "total_strength": total_strength,
        "total_credits": round(total_credits, 1),
        "behavioral": behavioral,
        "opinion": opinion,
        "avg_strength": round(total_strength / len(ev), 1) if ev else 0,
    }


def assumption_risk_report(team_id):
    """Flag high-importance untested assumptions — the ones that can kill a venture."""
    assums = db.list_assumptions(team_id)
    exposed = [a for a in assums
               if a["importance"] >= 4 and a["status"] in ("Untested", "Ignored")]
    return {
        "total": len(assums),
        "untested": sum(1 for a in assums if a["status"] == "Untested"),
        "supported": sum(1 for a in assums if a["status"] == "Supported"),
        "refuted": sum(1 for a in assums if a["status"] == "Refuted"),
        "exposed": exposed,
    }


# --------------------------------------------------------------------------- #
# Productive failure — mini-pivots (available from round 1) and evidence-driven
# course-correction counts. Changing your mind early, cheaply, is rewarded.
# --------------------------------------------------------------------------- #
MINI_PIVOT_CREDIT = 3   # Evidence Credits for logging a course-correction (learning reward)


def log_mini_pivot(team_id, original, evidence, change):
    """Record a lightweight, self-approved course-correction and reward the learning.

    Unlike the formal Pivot Petition (a late-round committee process), a mini-pivot
    is available from the very first round so productive failure is normalised early,
    when the cost of being wrong is low."""
    rnd = db.current_round()
    db.add_pivot(team_id,
                 {"original_assum": original, "challenge_evid": evidence,
                  "proposed_change": change},
                 kind="mini", status="Logged", round_no=rnd)
    db.adjust_resources(team_id, credits=MINI_PIVOT_CREDIT, kind="learning",
                        description="Mini-pivot logged — learning from evidence",
                        allow_negative=True)
    return MINI_PIVOT_CREDIT


def evidence_based_pivots(team_id):
    """How much the team has changed its mind based on evidence."""
    pivots = db.list_pivots(team_id)
    mini = sum(1 for p in pivots if (p.get("kind") or "formal") == "mini")
    approved = sum(1 for p in pivots
                   if (p.get("kind") or "formal") == "formal"
                   and p["status"] in ("Approved", "Conditional"))
    refuted = sum(1 for a in db.list_assumptions(team_id) if a["status"] == "Refuted")
    return {"mini": mini, "approved_formal": approved, "refuted": refuted,
            "total": mini + approved + refuted}


def sync_ai_logs(team_id):
    """Auto-verify AI logs whose linked test has resolved, so 'verification' is a real
    action the team already takes, not extra paperwork.

    A linked experiment/assumption that came back Supported verifies the AI's claim;
    Refuted rejects it. Only touches logs still 'Unverified'. Returns count changed."""
    exps = {e["id"]: e for e in db.list_experiments(team_id)}
    assums = {a["id"]: a for a in db.list_assumptions(team_id)}
    changed = 0
    for l in db.list_ai_logs(team_id):
        if l["status"] != "Unverified":
            continue
        outcome = None
        if l.get("experiment_id") and l["experiment_id"] in exps:
            outcome = exps[l["experiment_id"]]["outcome"]
        elif l.get("assumption_id") and l["assumption_id"] in assums:
            outcome = assums[l["assumption_id"]]["status"]
        if outcome == "Supported":
            db.update_ai_log(l["id"], status="Verified")
            changed += 1
        elif outcome == "Refuted":
            db.update_ai_log(l["id"], status="Rejected")
            changed += 1
    return changed


def ai_verification_rate(team_id):
    """Share of AI-assist logs the team actually EVALUATED (reached a decided status),
    rather than leaving as raw 'Unverified'. Rewards verification, not usage.

    Returns None when the team hasn't logged any AI use (so non-use isn't penalised)."""
    logs = db.list_ai_logs(team_id)
    if not logs:
        return None
    decided = sum(1 for l in logs if l["status"] in ("Verified", "Rejected", "Modified"))
    return decided / len(logs)


def ai_unverified_count(team_id):
    return sum(1 for l in db.list_ai_logs(team_id) if l["status"] == "Unverified")


# --------------------------------------------------------------------------- #
# Learning metrics — a per-round snapshot the team can watch trend over time.
# --------------------------------------------------------------------------- #
def learning_metrics(team_id):
    ev = evidence_summary(team_id)
    total_ev = ev["count"]
    assums = db.list_assumptions(team_id)
    important = [a for a in assums if a["importance"] >= COVERAGE_IMPORTANCE_MIN]
    tested = sum(1 for a in important if a["status"] in ("Supported", "Refuted"))
    piv = evidence_based_pivots(team_id)
    airate = ai_verification_rate(team_id)
    return {
        "behavioral": ev["behavioral"],
        "opinion": ev["opinion"],
        "evidence_items": total_ev,
        "behavioral_ratio": round(ev["behavioral"] / total_ev, 3) if total_ev else 0.0,
        "avg_strength": ev["avg_strength"],
        "test_coverage": round(tested / len(important), 3) if important else 0.0,
        "evidence_coverage": evidence_coverage(team_id),
        "pivots_evidence": piv["total"],
        "ai_verification": round(airate, 3) if airate is not None else None,
    }


def snapshot_metrics(team_id, rnd=None):
    rnd = db.current_round() if rnd is None else rnd
    db.save_metrics_snapshot(team_id, rnd, json.dumps(learning_metrics(team_id)))


def metrics_trend(team_id):
    """The team's learning metrics per round (stored history + a live point for the
    current round), oldest first."""
    out = []
    for h in db.list_metrics_history(team_id):
        try:
            m = json.loads(h["metrics"])
        except (ValueError, TypeError):
            m = {}
        m["round"] = h["round"]
        out.append(m)
    cur = db.current_round()
    if not any(o.get("round") == cur for o in out):
        m = learning_metrics(team_id)
        m["round"] = cur
        out.append(m)
    return sorted(out, key=lambda o: o["round"])


# --------------------------------------------------------------------------- #
# Engagement — team identity, badges/streaks/levels, and the narrative arc
# --------------------------------------------------------------------------- #
def team_identity(team):
    """Display name, colour, and mascot for a team (with sensible defaults)."""
    return {
        "display": (team.get("display_name") or team.get("name") or "Team"),
        "color": team.get("color") or content.DEFAULT_TEAM_COLOR,
        "mascot": team.get("mascot") or content.DEFAULT_TEAM_MASCOT,
    }


def commit_streak(team_id):
    """Consecutive most-recent rounds the team committed (an on-time streak)."""
    cur = db.current_round()
    streak = 0
    for r in range(cur, 0, -1):
        c = db.get_commitment(team_id, r)
        if c and c["committed"]:
            streak += 1
        else:
            break
    return streak


def founder_level(team_id):
    """A team 'founder level' that grows with skill (training + learning by doing).
    Returns (level, total_effective_skill)."""
    total = sum(effective_skill(team_id, k) for k in content.FOUNDER_SKILL_KEYS)
    return 1 + int(total // 6), total


def team_badges(team_id):
    """Set of badge codes the team has EARNED, computed from current state."""
    codes = set()
    ev = db.list_evidence(team_id)
    es = evidence_summary(team_id)
    if ev:
        codes.add("first_evidence")
    if any((e["strength"] or 0) >= 9 for e in ev):
        codes.add("paying_customer")
    if es["count"] >= 4 and es["behavioral"] / es["count"] >= 0.5:
        codes.add("behavior_beats_opinion")
    piv = evidence_based_pivots(team_id)
    if piv["refuted"] or any(a["status"] == "Refuted" for a in db.list_assumptions(team_id)):
        codes.add("killed_assumption")
    cal = calibration_summary(team_id)
    if cal["n"] >= 2 and cal["overconfidence_gap"] is not None \
            and abs(cal["overconfidence_gap"]) <= 10:
        codes.add("well_calibrated")
    ctypes = {c["ctype"] for c in db.list_canvases(team_id)}
    if {"customer_profile", "vpc", "bmc"} <= ctypes:
        codes.add("model_builder")
    if piv["mini"]:
        codes.add("course_corrector")
    if any(l["status"] == "Verified" for l in db.list_ai_logs(team_id)):
        codes.add("ai_auditor")
    if commit_streak(team_id) >= 3:
        codes.add("on_a_roll")
    if evidence_coverage(team_id) >= 0.6:
        codes.add("evidence_machine")
    return codes


def sync_badges(team_id):
    """Award any newly-earned badges; return the set earned THIS check (for a toast)."""
    earned = db.get_earned_badges(team_id)
    current = team_badges(team_id)
    new = current - earned
    for c in new:
        db.award_badge(team_id, c)
    return new


def badge_progress(team_id):
    """All badges with earned/locked state, for the trophy case."""
    earned = team_badges(team_id)
    out = []
    for code, name, emoji, desc in content.BADGES:
        out.append({"code": code, "name": name, "emoji": emoji, "desc": desc,
                    "earned": code in earned})
    return out


# ---- Narrative arc + investor persona -------------------------------------- #
def story_phase(rnd=None):
    rnd = db.current_round() if rnd is None else rnd
    return content.narrative_phase(rnd, total_rounds())


def investor_line(team_id):
    """Vera Sloan's in-character nudge, chosen by how the team is actually doing."""
    es = evidence_summary(team_id)
    risk = assumption_risk_report(team_id)
    piv = evidence_based_pivots(team_id)
    cov = evidence_coverage(team_id)
    lines = content.INVESTOR_LINES
    if cov >= 0.6:
        return lines["pitch_ready"]
    if risk["exposed"]:
        return lines["risky_untested"]
    if es["count"] == 0:
        return lines["no_evidence"]
    if es["count"] >= 3 and es["opinion"] > es["behavioral"]:
        return lines["opinion_heavy"]
    if piv["total"]:
        return lines["pivoted"]
    if es["behavioral"] >= 2:
        return lines["good_evidence"]
    return lines["no_evidence"]


def story_event_text(category, exposes):
    """Wrap a market event in the narrative voice while still naming the assumption."""
    intro = content.EVENT_STORY_INTRO.get(category, "A twist in the story:")
    tail = f" It puts pressure on your assumption: “{exposes}.”" if exposes else ""
    return intro, tail


# ---- Cohort leaderboard ---------------------------------------------------- #
LEADERBOARD_METRICS = {
    "round_score": "Round score",
    "valuation": "Venture valuation",
    "coverage": "Evidence coverage",
}


def leaderboard(game_id=None, metric="round_score"):
    """Ranked standings for a game's teams by the chosen metric (highest first)."""
    gid = game_id if game_id is not None else db.active_game_id()
    rnd = db.current_round()
    rows = []
    for t in db.list_teams(gid):
        rs = round_score(t["id"], rnd)
        val = compute_valuation(t["id"])
        rows.append({
            "team": t, "round_score": rs["score"], "valuation": val["valuation"],
            "coverage": val["evidence_coverage"], "badges": len(team_badges(t["id"])),
        })
    key = metric if metric in ("round_score", "valuation", "coverage") else "round_score"
    rows.sort(key=lambda r: r[key], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


# ---- Demo Day / Evidence Exchange ------------------------------------------ #
DEMO_VOTES_PER_TEAM = 3


def demo_is_open(game_id=None):
    gid = game_id if game_id is not None else db.active_game_id()
    return db.get_setting(f"demo_open:{gid}", "0") == "1"


def set_demo_open(game_id, on):
    db.set_setting(f"demo_open:{game_id}", "1" if on else "0")


def strongest_evidence(team_id):
    ev = db.list_evidence(team_id)
    return max(ev, key=lambda e: e["strength"]) if ev else None


def demo_results(game_id=None):
    """Peer-vote tally joined to teams, ranked."""
    gid = game_id if game_id is not None else db.active_game_id()
    tally = db.vote_tally(gid)
    rows = [{"team": t, "votes": tally.get(t["id"], 0)} for t in db.list_teams(gid)]
    rows.sort(key=lambda r: r["votes"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


# ---- Spaced retrieval ------------------------------------------------------ #
def spaced_review_concepts(rnd, n=2):
    """1–2 concepts from EARLIER rounds (with a quiz) to re-test for retrieval practice."""
    if rnd <= 1:
        return []
    pool = []
    for r in range(1, rnd):
        for c in round_concepts(r):
            if content.CONCEPT_QUIZ.get(c) and c not in pool:
                pool.append(c)
    if not pool:
        return []
    picks = []
    for i in range(min(n, len(pool))):
        picks.append(pool[(rnd * 7 + i * 3) % len(pool)])
    return list(dict.fromkeys(picks))


# ---- Instructor misconception radar ---------------------------------------- #
def misconception_report(game_id=None):
    """Which concepts teams get WRONG most (from stored true/false), plus how many
    important assumptions remain untested across the cohort."""
    gid = game_id if game_id is not None else db.active_game_id()
    teams = db.list_teams(gid)
    stats = {}
    for t in teams:
        for r in range(1, total_rounds() + 1):
            answers = db.get_round_answers(t["id"], r)
            for c, raw in answers.items():
                if not content.CONCEPT_QUIZ.get(c):
                    continue
                parsed = _parse_concept_response(raw)
                if parsed["quiz"] is None:
                    continue
                s = stats.setdefault(c, {"answered": 0, "wrong": 0})
                s["answered"] += 1
                if not concept_quiz_correct(c, parsed["quiz"]):
                    s["wrong"] += 1
    concepts = [{"concept": c, "answered": s["answered"], "wrong": s["wrong"],
                 "wrong_rate": (s["wrong"] / s["answered"] if s["answered"] else 0)}
                for c, s in stats.items()]
    concepts.sort(key=lambda x: (x["wrong_rate"], x["wrong"]), reverse=True)
    total_imp = untested = 0
    exposed_by_concept = {}
    for t in teams:
        for a in db.list_assumptions(t["id"]):
            if a["importance"] >= COVERAGE_IMPORTANCE_MIN:
                total_imp += 1
                if a["status"] in ("Untested", "Ignored"):
                    untested += 1
    return {"concepts": concepts, "teams": len(teams),
            "total_important": total_imp, "untested_important": untested}


# --------------------------------------------------------------------------- #
# Optional AI comment (bring-your-own-key) — deterministic feedback is always on;
# this ADDS a richer LLM comment only when an instructor supplies a key. It uses
# the standard library (no extra dependency) and degrades to None on any problem.
# --------------------------------------------------------------------------- #
AI_PROVIDERS = {
    "groq": {"label": "Groq — Llama 3.1 (fast, free tier)", "kind": "openai",
             "base": "https://api.groq.com/openai/v1", "model": "llama-3.1-8b-instant",
             "keys_url": "https://console.groq.com/keys"},
    "gemini": {"label": "Google Gemini (free tier)", "kind": "gemini",
               "base": "https://generativelanguage.googleapis.com/v1beta",
               "model": "gemini-2.5-flash", "keys_url": "https://aistudio.google.com/apikey"},
    "openai": {"label": "OpenAI-compatible (custom base URL)", "kind": "openai",
               "base": "https://api.openai.com/v1", "model": "gpt-4o-mini",
               "keys_url": "https://platform.openai.com/api-keys"},
}


def get_ai_config():
    prov = db.get_setting("ai_provider", "groq")
    if prov not in AI_PROVIDERS:
        prov = "groq"
    p = AI_PROVIDERS[prov]
    return {
        "provider": prov,
        "enabled": db.get_setting("ai_enabled", "0") == "1",
        "key": db.get_setting("ai_key", "") or "",
        "model": db.get_setting("ai_model", "") or p["model"],
        "base": db.get_setting("ai_base", "") or p["base"],
        "kind": p["kind"],
    }


def set_ai_config(provider=None, enabled=None, key=None, model=None, base=None):
    if provider is not None:
        db.set_setting("ai_provider", provider)
    if enabled is not None:
        db.set_setting("ai_enabled", "1" if enabled else "0")
    if key is not None:
        db.set_setting("ai_key", key)
    if model is not None:
        db.set_setting("ai_model", model)
    if base is not None:
        db.set_setting("ai_base", base)


def ai_available():
    cfg = get_ai_config()
    return bool(cfg["enabled"] and cfg["key"])


def _http_json(url, payload, headers, timeout=25):
    import urllib.request
    import urllib.error
    import json as _json
    data = _json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return _json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", "replace")[:300]
        except Exception:
            pass
        raise RuntimeError(f"HTTP {e.code} — {body or e.reason}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"Network error — {getattr(e, 'reason', e)}")


def ai_comment(prompt, system=None, cfg=None, timeout=25, return_error=False):
    """Return an LLM comment, or None if AI isn't configured or anything fails.

    Deterministic feedback never depends on this — it's purely additive. When
    return_error=True, returns (text_or_None, error_message_or_None)."""
    cfg = cfg or get_ai_config()

    def _out(text, err):
        return (text, err) if return_error else text

    if not (cfg["enabled"] and cfg["key"]):
        return _out(None, "AI is off or no API key is set.")
    try:
        if cfg["kind"] == "gemini":
            url = f"{cfg['base'].rstrip('/')}/models/{cfg['model']}:generateContent"
            sys_txt = (system + "\n\n") if system else ""
            body = {"contents": [{"parts": [{"text": sys_txt + prompt}]}],
                    "generationConfig": {"maxOutputTokens": 300, "temperature": 0.6}}
            headers = {"Content-Type": "application/json", "x-goog-api-key": cfg["key"]}
            out = _http_json(url, body, headers, timeout)
            cands = out.get("candidates") or []
            if not cands:
                return _out(None, f"Model returned no text ({str(out)[:150]}).")
            parts = cands[0].get("content", {}).get("parts", [])
            text = "".join(p.get("text", "") for p in parts).strip()
            return _out(text or None, None if text else "Empty response from model.")
        # OpenAI-compatible (Groq / OpenAI / others)
        url = f"{cfg['base'].rstrip('/')}/chat/completions"
        msgs = ([{"role": "system", "content": system}] if system else []) \
            + [{"role": "user", "content": prompt}]
        body = {"model": cfg["model"], "messages": msgs,
                "max_tokens": 300, "temperature": 0.6}
        headers = {"Content-Type": "application/json",
                   "Authorization": f"Bearer {cfg['key']}"}
        out = _http_json(url, body, headers, timeout)
        return _out(out["choices"][0]["message"]["content"].strip(), None)
    except Exception as e:   # noqa: BLE001 — any failure degrades gracefully
        return _out(None, str(e))


def ai_test_key(cfg=None):
    """Quick round-trip to validate the key. Returns (ok, message) with the real error."""
    text, err = ai_comment("Reply with the single word: OK", cfg=cfg, timeout=20,
                           return_error=True)
    if text:
        return True, f"Working ✓ (model replied: “{text[:40]}”)"
    return False, err or "No response — check the key, model, and provider."


def ai_round_comment(team_id, round_no=None):
    """A short, in-character coaching comment grounded in the team's real state.
    Returns None if AI isn't configured."""
    if not ai_available():
        return None
    team = db.get_team(team_id)
    rnd = round_no or db.current_round()
    es = evidence_summary(team_id)
    risk = assumption_risk_report(team_id)
    val = compute_valuation(team_id)
    _pi, _pn, _pe, _pt = content.narrative_phase(rnd, total_rounds())
    facts = (f"Round {rnd} ({_pn} phase). Evidence items: {es['count']} "
             f"({es['behavioral']} behavioral, {es['opinion']} opinion, "
             f"avg strength {es['avg_strength']}/10). "
             f"Important untested assumptions: {len(risk['exposed'])}. "
             f"Evidence coverage: {val['evidence_coverage']*100:.0f}%. "
             f"Territory: {team['opportunity']}.")
    system = (f"You are {content.INVESTOR['name']}, {content.INVESTOR['title']} — sharp, warm, "
              "allergic to hype. Give a startup team 2–3 sentences of specific, honest coaching "
              "for this round. Reward evidence and behavior over opinion; name the single most "
              "useful next move. No lists, no preamble, under 70 words.")
    return ai_comment(facts, system=system)


def quick_setup_teams(n_teams, difficulty, opportunity_mode="distinct",
                      opportunity_choice=None, founder_mode="balanced",
                      name_prefix="Team", clear_existing=False):
    """Create a balanced cohort of teams in one step.

    Every team receives IDENTICAL starting resources and market potential (from the
    chosen difficulty preset), which is what equalizes their odds of success. Only
    flavor differs by option:
      opportunity_mode : "same"     -> all teams get opportunity_choice
                         "distinct" -> rotate distinct territories (still balanced)
      founder_mode     : "balanced" -> everyone gets the neutral balanced card
                         "varied"   -> rotate founder archetypes, but resources are
                                       still forced equal so no one is advantaged

    Returns a list of {name, code, territory} for the teams created.
    """
    preset = content.DIFFICULTY_LEVELS.get(difficulty, content.DIFFICULTY_LEVELS["Standard"])
    if clear_existing:
        for t in db.list_teams():
            db.delete_team(t["id"])
    db.set_setting("difficulty", difficulty)

    territories = content.OPPORTUNITY_TERRITORIES
    if opportunity_mode == "same" and not opportunity_choice:
        opportunity_choice = territories[0]

    created = []
    for i in range(int(n_teams)):
        territory = (opportunity_choice if opportunity_mode == "same"
                     else territories[i % len(territories)])
        if founder_mode == "varied":
            base_card = dict(content.FOUNDER_CARDS[i % len(content.FOUNDER_CARDS)])
        else:
            base_card = dict(content.BALANCED_FOUNDER_CARD)
        # Force equal resources regardless of the card's own defaults => fair start.
        base_card["budget"] = preset["capital"]
        base_card["hours"] = preset["hours"]

        name = f"{name_prefix} {i + 1}"
        code = db.create_team(
            name, territory, base_card,
            capital=preset["capital"], evidence_credits=preset["credits"],
            founder_hours=preset["hours"], market_potential=preset["market_potential"],
            hours_per_round=preset["hours"],   # recommended weekly hours
        )
        # seed a default business-dev budget (effort starts near the green line)
        new_team = db.get_team_by_code(code)
        b = default_build(db.get_team(new_team["id"]))
        db.update_team(new_team["id"], build_budget=b, founder_hours=b)
        send_welcome(new_team["id"])   # Round-1 welcome + subtle hints in the Inbox
        created.append({"name": name, "code": code, "territory": territory})
    return created


# --------------------------------------------------------------------------- #
# Roster import — build games and teams from a class contact list (xlsx/csv).
#   A row is a student. Columns (case-insensitive, flexible):
#     Class, Section  -> which GAME the team belongs to
#     Team            -> which team within that game
#     FirstName, LastName, PrimaryEmail (or Email) -> the member
# --------------------------------------------------------------------------- #
_ROSTER_ALIASES = {
    "class": "class", "section": "section", "team": "team",
    "firstname": "first", "first name": "first", "first": "first",
    "lastname": "last", "last name": "last", "last": "last",
    "primaryemail": "email", "email": "email", "e-mail": "email",
}


def read_table(file_bytes, filename):
    """Return (headers, rows-as-dicts) from an .xlsx or .csv upload."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        import csv
        import io
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [r for r in reader if any((c or "").strip() for c in r)]
        if not rows:
            return [], []
        headers = [str(h or "").strip() for h in rows[0]]
        data = [dict(zip(headers, r)) for r in rows[1:]]
        return headers, data
    # xlsx
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        data.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return headers, data


def _norm_row(row):
    """Map a raw row's columns onto normalized keys via the aliases."""
    out = {}
    for k, v in row.items():
        key = _ROSTER_ALIASES.get(str(k).strip().lower())
        if key:
            out[key] = v
    return out


def parse_roster(headers, rows):
    """Group rows into games (Class + Section) → teams (Team #) → members.

    Returns a list of {"game": name, "teams": [{"label", "members":[{name,email}]}]}
    and a list of warnings."""
    warnings = []
    games = {}      # game_name -> {team_label -> [members]}
    order = []      # preserve game order
    for raw in rows:
        r = _norm_row(raw)
        cls = str(r.get("class") or "").strip()
        section = str(r.get("section") or "").strip()
        team = str(r.get("team") or "").strip()
        first = str(r.get("first") or "").strip()
        last = str(r.get("last") or "").strip()
        email = str(r.get("email") or "").strip()
        name = (first + " " + last).strip() or email
        if not (cls or section) or not team:
            warnings.append(f"Skipped a row missing Class/Section or Team: {name or raw}")
            continue
        game_name = cls if not section else f"{cls} · Section {section}"
        game_name = game_name.strip(" ·")
        if game_name not in games:
            games[game_name] = {}
            order.append(game_name)
        label = f"Team {team}"
        games[game_name].setdefault(label, [])
        if name:
            games[game_name][label].append({"name": name, "email": email})
    result = []
    for gname in order:
        teams = [{"label": lbl, "members": mem}
                 for lbl, mem in sorted(games[gname].items(),
                                        key=lambda kv: _team_sort_key(kv[0]))]
        result.append({"game": gname, "teams": teams})
    return result, warnings


def _team_sort_key(label):
    digits = "".join(ch for ch in label if ch.isdigit())
    return (int(digits) if digits else 9999, label)


def import_roster(parsed, difficulty="Standard", opportunity_mode="distinct",
                  founder_mode="balanced", merge_into_game_id=None):
    """Create games and teams from a parsed roster. Returns a summary dict."""
    preset = content.DIFFICULTY_LEVELS.get(difficulty, content.DIFFICULTY_LEVELS["Standard"])
    db.set_setting("difficulty", difficulty)
    territories = content.OPPORTUNITY_TERRITORIES
    summary = {"games": 0, "teams": 0, "members": 0, "details": []}

    for gi, g in enumerate(parsed):
        if merge_into_game_id and len(parsed) == 1:
            gid = merge_into_game_id
        else:
            gid = db.create_game(g["game"])
            summary["games"] += 1
        db.set_active_game(gid)
        made = []
        for ti, t in enumerate(g["teams"]):
            territory = (territories[0] if opportunity_mode == "same"
                         else territories[ti % len(territories)])
            if founder_mode == "varied":
                base_card = dict(content.FOUNDER_CARDS[ti % len(content.FOUNDER_CARDS)])
            else:
                base_card = dict(content.BALANCED_FOUNDER_CARD)
            base_card["budget"] = preset["capital"]
            base_card["hours"] = preset["hours"]
            code = db.create_team(
                t["label"], territory, base_card,
                capital=preset["capital"], evidence_credits=preset["credits"],
                founder_hours=preset["hours"], market_potential=preset["market_potential"],
                hours_per_round=preset["hours"], game_id=gid, roster=t["members"],
            )
            nt = db.get_team_by_code(code)
            b = default_build(db.get_team(nt["id"]))
            db.update_team(nt["id"], build_budget=b, founder_hours=b)
            send_welcome(nt["id"])
            made.append({"team": t["label"], "code": code, "members": len(t["members"])})
            summary["teams"] += 1
            summary["members"] += len(t["members"])
        summary["details"].append({"game": g["game"], "game_id": gid, "teams": made})
    return summary


def roster_template_workbook():
    """Build an example roster .xlsx (openpyxl Workbook) with the expected columns."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roster"
    headers = ["FirstName", "LastName", "PrimaryEmail", "Class", "Section", "Team",
               "Team Member 1", "Team Member 2", "Team Member 3", "Team Member 4"]
    ws.append(headers)
    example = [
        ("Ada", "Lovelace", "ada@example.edu", "MGT 301", 1, 1),
        ("Alan", "Turing", "alan@example.edu", "MGT 301", 1, 1),
        ("Grace", "Hopper", "grace@example.edu", "MGT 301", 1, 2),
        ("Katherine", "Johnson", "kj@example.edu", "MGT 301", 1, 2),
        ("Mae", "Jemison", "mae@example.edu", "MGT 301", 2, 1),
        ("Sally", "Ride", "sally@example.edu", "MGT 301", 2, 1),
    ]
    # Fill the redundant "Team Member N" columns per team for readability.
    from collections import defaultdict
    teams = defaultdict(list)
    for f, l, e, c, s, tm in example:
        teams[(c, s, tm)].append(f"{f} {l}")
    for f, l, e, c, s, tm in example:
        members = teams[(c, s, tm)]
        row = [f, l, e, c, s, tm] + members[:4] + [""] * (4 - len(members[:4]))
        ws.append(row)
    for col in "ABCDEFGHIJ":
        ws.column_dimensions[col].width = 16
    return wb


def cohort_balance(team_list):
    """Report how equal the teams' starting resources are (fairness check).

    Returns a dict with the spread of capital, credits, hours, and market potential.
    A spread of 0 means a perfectly balanced cohort.
    """
    if not team_list:
        return None

    def spread(key):
        vals = [t[key] for t in team_list]
        return {"min": min(vals), "max": max(vals), "spread": round(max(vals) - min(vals), 2)}

    balanced = all(
        spread(k)["spread"] == 0
        for k in ("capital", "evidence_credits", "founder_hours", "market_potential")
    )
    return {
        "balanced": balanced,
        "capital": spread("capital"),
        "credits": spread("evidence_credits"),
        "hours": spread("founder_hours"),
        "market_potential": spread("market_potential"),
    }


def _alignment(props, key="tokens"):
    """Tokens-weighted average evidence support (0..1) for an allocation."""
    total = sum(max(0, p[key] or 0) for p in props)
    if total <= 0:
        return None  # no tokens allocated => undefined
    weighted = sum((p[key] or 0) * (p["evidence_strength"] or 0) / 10.0 for p in props)
    return weighted / total


def preview_vp_auction(team_id, allocations):
    """Compute (without committing) the auction outcome for a proposed allocation.

    allocations: {vp_id: tokens}. Returns a dict describing tax, dividend, net,
    and per-VP breakdown so the student can see consequences before submitting.
    """
    props = db.list_value_props(team_id)
    if not props:
        return {"ok": False, "reason": "No value propositions defined."}
    if len(props) < content.MIN_VALUE_PROPS:
        return {"ok": False,
                "reason": f"Field at least {content.MIN_VALUE_PROPS} value propositions "
                          f"(you have {len(props)})."}
    total = sum(max(0, int(allocations.get(p["id"], 0))) for p in props)
    if total <= 0:
        return {"ok": False, "reason": "Allocate at least some Venture Tokens."}
    if total > content.VENTURE_TOKEN_POOL:
        return {"ok": False,
                "reason": f"Allocation ({total}) exceeds the {content.VENTURE_TOKEN_POOL}-token pool."}

    proposed = [{**p, "tokens": int(allocations.get(p["id"], 0))} for p in props]
    alignment = _alignment(proposed, "tokens")

    # The "previous" allocation is the team's currently committed one (in `tokens`).
    # It is undefined (None) until the team has run at least one prior auction.
    prev_total = sum(max(0, p["tokens"] or 0) for p in props)
    prev_alignment = _alignment(props, "tokens") if prev_total > 0 else None

    tax = round((1 - alignment) * content.OVERCONFIDENCE_TAX_MAX, 1)
    if prev_alignment is None:
        dividend = 0.0
    else:
        dividend = round(max(0.0, alignment - prev_alignment) * content.LEARNING_DIVIDEND_MAX, 1)
    net = round(dividend - tax, 1)

    breakdown = []
    for p in proposed:
        support = (p["evidence_strength"] or 0) / 10.0
        breakdown.append({
            "id": p["id"], "name": p["name"], "tokens": p["tokens"],
            "evidence_strength": p["evidence_strength"] or 0,
            "unsupported_tokens": round(p["tokens"] * (1 - support), 1),
        })
    return {
        "ok": True, "total_tokens": total, "alignment": round(alignment, 3),
        "prev_alignment": None if prev_alignment is None else round(prev_alignment, 3),
        "tax": tax, "dividend": dividend, "net": net, "breakdown": breakdown,
    }


def run_vp_auction(team_id, allocations, round_no):
    """Commit an auction round: save allocations, apply the net credit change,
    snapshot prev_tokens, and record the result. Returns the preview dict plus
    a committed flag.
    """
    result = preview_vp_auction(team_id, allocations)
    if not result.get("ok"):
        return result

    props = db.list_value_props(team_id)
    # Snapshot current tokens as prev, then store the new allocation.
    for p in props:
        new_tokens = int(allocations.get(p["id"], 0))
        db.update_value_prop(p["id"], prev_tokens=p["tokens"], tokens=new_tokens)

    note = (f"Alignment {result['alignment']} "
            f"(prev {result['prev_alignment']}); "
            f"tax {result['tax']}, dividend {result['dividend']}")
    db.record_vp_result(
        team_id, round_no, result["total_tokens"], result["alignment"],
        result["prev_alignment"], result["tax"], result["dividend"],
        result["net"], note,
    )
    if result["net"] != 0:
        db.adjust_resources(
            team_id, credits=result["net"], kind="vp_auction",
            description=note, allow_negative=True,
        )
    result["committed"] = True
    return result


def experiment_efficiency(team_id):
    """Learning (evidence strength) generated per dollar and founder-hour."""
    exps = db.list_experiments(team_id)
    strength = sum(e["evidence_strength"] for e in exps
                   if e["outcome"] in ("Supported", "Refuted"))
    money = sum(e["cost_money"] for e in exps)
    hours = sum(e["cost_time"] for e in exps)
    return {
        "experiments": len(exps),
        "resolved": sum(1 for e in exps if e["outcome"] in ("Supported", "Refuted")),
        "strength_per_dollar": round(strength / money, 3) if money else 0,
        "strength_per_hour": round(strength / hours, 3) if hours else 0,
    }


# ---- Calibration: did the team's forecasts match reality? ------------------ #
def prediction_correct(exp):
    """Whether a resolved experiment's outcome matched the team's prediction.
    Returns True/False, or None if not comparable (no prediction or not resolved)."""
    pred = (exp.get("predicted_outcome") or "").strip()
    actual = exp.get("outcome")
    if not pred or actual not in ("Supported", "Refuted", "Inconclusive"):
        return None
    return pred == actual


def calibration_summary(team_id):
    """Compare the team's up-front predictions with what actually happened.

    Overconfidence gap = average stated confidence − actual hit-rate. Positive means
    the team was more sure than it should have been — the lesson the loop teaches."""
    exps = db.list_experiments(team_id)
    scored = [e for e in exps if prediction_correct(e) is not None
              and e.get("confidence") is not None]
    if not scored:
        return {"n": 0, "correct": 0, "hit_rate": None, "avg_confidence": None,
                "overconfidence_gap": None}
    correct = sum(1 for e in scored if prediction_correct(e))
    hit = 100.0 * correct / len(scored)
    avg_conf = sum(float(e["confidence"]) for e in scored) / len(scored)
    return {"n": len(scored), "correct": correct, "hit_rate": round(hit, 0),
            "avg_confidence": round(avg_conf, 0),
            "overconfidence_gap": round(avg_conf - hit, 0)}


# --------------------------------------------------------------------------- #
# Auto-Director (autopilot)
#
# Derives the Director's per-round decisions from each team's actual round input
# (their canvases, evidence, experiments, assumptions, auction, AI logs, etc.):
#   • predicted dashboard scores (0–100) per dimension
#   • a suggested market event aimed at the team's biggest exposed risk
#   • a recommended decision for each pending pivot petition
# Every suggestion is just that — the Director can override before applying.
# --------------------------------------------------------------------------- #
def _clamp100(x):
    return int(max(0, min(100, round(x))))


def _filled_blocks(canvas):
    if not canvas:
        return 0, 0
    data = canvas.get("data", {}) or {}
    total = len(data)
    filled = sum(1 for v in data.values() if v and str(v).strip())
    return filled, total


def default_score_weights():
    """Per-dimension weight multipliers, default 1.0 for every dimension."""
    return {dim: 1.0 for dim in content.DIMENSION_NAMES}


def get_score_weights():
    """Load the Director's tunable per-dimension weights (1.0 = default)."""
    raw = db.get_setting("score_weights")
    weights = default_score_weights()
    if raw:
        try:
            saved = json.loads(raw)
            for dim in weights:
                if dim in saved:
                    weights[dim] = float(saved[dim])
        except (ValueError, TypeError):
            pass
    return weights


def set_score_weights(weights):
    clean = {dim: round(float(weights.get(dim, 1.0)), 2) for dim in content.DIMENSION_NAMES}
    db.set_setting("score_weights", json.dumps(clean))


# --------------------------------------------------------------------------- #
# Round score (0–100) — a single grade for the work a team committed this round.
#   Four merit components (each 0–100), blended by Director-tunable weights:
#     • Commitment  — how much of the round's required work is complete.
#     • Evidence    — strength/quality of the evidence gathered.
#     • Coherence   — business-model + value-proposition coherence.
#     • Concepts    — this round's concept-checks answered.
#   A risk penalty subtracts for important, still-untested assumptions.
#   A strictness dial then stretches (strict) or lifts (lenient) the result.
# --------------------------------------------------------------------------- #
ROUND_SCORE_COMPONENTS = ["commitment", "evidence", "coherence", "concepts",
                          "ai_verification"]
DEFAULT_ROUND_SCORE_WEIGHTS = {"commitment": 35, "evidence": 25, "coherence": 15,
                               "concepts": 15, "ai_verification": 10}
RISK_PENALTY_PER_ITEM = 5     # points lost per important untested assumption
RISK_PENALTY_CAP = 25


def default_round_score_config():
    return {"weights": dict(DEFAULT_ROUND_SCORE_WEIGHTS), "strictness": 50}


def get_round_score_config():
    """Director's round-score settings: component weights (0–100) + strictness (0–100)."""
    cfg = default_round_score_config()
    raw = db.get_setting("round_score_config")
    if raw:
        try:
            saved = json.loads(raw)
            for k in cfg["weights"]:
                if k in saved.get("weights", {}):
                    cfg["weights"][k] = float(saved["weights"][k])
            if "strictness" in saved:
                cfg["strictness"] = float(saved["strictness"])
        except (ValueError, TypeError):
            pass
    return cfg


def set_round_score_config(weights, strictness):
    cfg = {"weights": {k: round(float(weights.get(k, DEFAULT_ROUND_SCORE_WEIGHTS[k])), 1)
                       for k in ROUND_SCORE_COMPONENTS},
           "strictness": round(float(strictness), 1)}
    db.set_setting("round_score_config", json.dumps(cfg))


def _strictness_gamma(strictness):
    """Map a 0–100 strictness dial to a gamma exponent for the scoring curve.

    50 = neutral (gamma 1.0). Above 50 makes high scores harder to reach
    (gamma > 1); below 50 is more forgiving (gamma < 1)."""
    import math
    s = max(0.0, min(100.0, float(strictness)))
    return math.exp((s - 50.0) / 50.0 * 0.9)   # ~0.41 (lenient) .. 1.0 .. 2.46 (strict)


_ALIGN_STOPWORDS = set(
    "the a an and or of to for in on with your you our we are is it that this "
    "their they will can could would should about into from over under your "
    "customers customer business venture idea using make made need needs when "
    "what which where whom whose have will value team round".split())


def _context_terms(team_id):
    """Distinctive words from the team's territory + their own candidate ventures.

    These represent 'their business' — an answer that references them is grounded
    in the venture they're actually building, not generic boilerplate."""
    team = db.get_team(team_id)
    terms = set()
    if team:
        for w in re.findall(r"[A-Za-z]{4,}", team["opportunity"] or ""):
            terms.add(w.lower())
    for v in db.get_ventures(team_id):
        for field in (v.get("name", ""), v.get("notes", "")):
            for w in re.findall(r"[A-Za-z]{4,}", field or ""):
                terms.add(w.lower())
    return terms - _ALIGN_STOPWORDS


def _answer_aligned(answer, terms):
    if not answer or not terms:
        return False
    words = {w.lower() for w in re.findall(r"[A-Za-z]{4,}", answer)}
    return bool(words & terms)


def _concepts_component(team_id, rnd):
    """This round's concept coverage, crediting answers that ALIGN with the team's
    territory/venture more than generic ones.

    Each of this round's concepts scores: 1.0 if answered and aligned to their
    business, 0.6 if answered but generic, 0 if unanswered. Returns (score0_100,
    answered, aligned, total)."""
    cp = concept_progress(team_id, rnd)
    if not cp:
        return 100.0, 0, 0, 0
    answers = db.get_round_answers(team_id, rnd)
    total = len(cp)
    score_sum = 0.0
    answered = aligned = 0
    for c in cp:
        if not c.get("needs_question", True):
            # Concept demonstrated by a decision — full credit once that decision is done.
            if c["done"]:
                score_sum += 1.0
            continue
        stt = concept_answer_status(team_id, rnd, c["concept"], answers)
        if stt["done"]:
            score_sum += 1.0                    # complete, quality answer + correct quiz
            answered += 1
            if stt["quality"]["checks"]["relevant"]:
                aligned += 1
        elif stt["text"].strip() or stt["quiz"] is not None:
            score_sum += 0.4                    # attempted but incomplete / wrong quiz
            answered += 1
    return 100.0 * score_sum / total, answered, aligned, total


def round_score_available(rnd):
    """Which score components are fair to grade in a given round — i.e. the tools
    they depend on have been introduced. Commitment and concepts are always fair
    (they concern this round's own tasks); evidence and coherence only once their
    tools exist, so teams are never marked down for things they couldn't do yet.
    AI verification is a base tool (always available) but only counts once a team
    has actually logged AI use — see round_score."""
    avail = {"commitment": True, "concepts": True, "ai_verification": True,
             "evidence": page_unlock_round("Evidence Ledger") <= rnd,
             "coherence": (canvas_unlock_round("vpc") <= rnd
                           or canvas_unlock_round("bmc") <= rnd)}
    return avail


def round_score(team_id, rnd, config=None):
    """Return a 0–100 round score with its component breakdown.

    Only counts what the team could actually DO this round: commitment (this round's
    decisions) and concept coverage always count; evidence and business-model
    coherence count only once their tools are introduced. Concept answers that align
    with the team's territory/venture are credited more than generic ones. Scores the
    committed snapshot if committed, else the team's current state."""
    cfg = config or get_round_score_config()
    weights = cfg["weights"]
    avail = round_score_available(rnd)

    # --- Commitment completion (from committed snapshot if present) ----------
    cstate = commitment_state(team_id, rnd)
    row = db.get_commitment(team_id, rnd)
    snap = None
    if cstate["committed"] and row and row.get("snapshot"):
        try:
            snap = json.loads(row["snapshot"])
        except (ValueError, TypeError):
            snap = None
    if snap is None:
        snap = commitment_snapshot(team_id, rnd)
    commitment = 100.0 * (snap["done"] / snap["total"]) if snap["total"] else 100.0

    # --- Concept coverage this round (alignment-weighted) --------------------
    concepts, answered, aligned, concept_total = _concepts_component(team_id, rnd)

    # --- Evidence & coherence — only from tools available by this round ------
    raw = _raw_scores(team_id)
    comps = {"commitment": commitment, "concepts": concepts}
    if avail["evidence"]:
        comps["evidence"] = _clamp100(raw.get("Evidence Strength", 0))
    if avail["coherence"]:
        parts = []
        if canvas_unlock_round("vpc") <= rnd:
            parts.append(_clamp100(raw.get("Value Proposition Fit", 0)))
        if canvas_unlock_round("bmc") <= rnd:
            parts.append(_clamp100(raw.get("Business-Model Coherence", 0)))
        if parts:
            comps["coherence"] = sum(parts) / len(parts)

    # --- AI verification — rewards evaluating AI, only if they've used AI ------
    airate = ai_verification_rate(team_id)
    if airate is not None:
        comps["ai_verification"] = 100.0 * airate

    # --- Weighted blend over AVAILABLE components only (renormalized) ---------
    counted = [k for k in ROUND_SCORE_COMPONENTS if k in comps]
    wsum = sum(weights.get(k, 0) for k in counted) or 1.0
    base = sum(comps[k] * weights.get(k, 0) for k in counted) / wsum

    # --- Strictness curve -----------------------------------------------------
    gamma = _strictness_gamma(cfg["strictness"])
    curved = 100.0 * (max(0.0, base) / 100.0) ** gamma

    # --- Risk penalty (only once assumptions are a thing this round) ----------
    if page_unlock_round("Assumption Map") <= rnd:
        exposed = len(assumption_risk_report(team_id)["exposed"])
    else:
        exposed = 0
    penalty = min(RISK_PENALTY_CAP, exposed * RISK_PENALTY_PER_ITEM)

    final = max(0.0, min(100.0, curved - penalty))
    return {
        # None for a component means "not available this round — not counted".
        "components": {k: (round(comps[k], 1) if k in comps else None)
                       for k in ROUND_SCORE_COMPONENTS},
        "counted": counted,
        "weights": {k: weights.get(k, 0) for k in ROUND_SCORE_COMPONENTS},
        "concept_answered": answered,
        "concept_aligned": aligned,
        "concept_total": concept_total,
        "base": round(base, 1),
        "curved": round(curved, 1),
        "penalty": penalty,
        "exposed_assumptions": exposed,
        "committed": cstate["committed"],
        "strictness": cfg["strictness"],
        "gamma": round(gamma, 3),
        "score": round(final, 1),
        "grade": round_score_band(final),
    }


def round_score_band(score):
    if score >= 90:
        return "Exceptional"
    if score >= 75:
        return "Strong"
    if score >= 60:
        return "Solid"
    if score >= 45:
        return "Developing"
    return "Needs work"


def _raw_scores(team_id):
    """Compute the UNWEIGHTED heuristic score for each dimension (0..~100 floats)."""
    ev = evidence_summary(team_id)
    risk = assumption_risk_report(team_id)
    eff = experiment_efficiency(team_id)

    cp = db.latest_canvas(team_id, "customer_profile")
    vpc = db.latest_canvas(team_id, "vpc")
    bmc = db.latest_canvas(team_id, "bmc")
    cp_filled, _ = _filled_blocks(cp)
    vpc_filled, _ = _filled_blocks(vpc)
    bmc_filled, _ = _filled_blocks(bmc)

    cp_versions = len(db.list_canvases(team_id, "customer_profile"))
    all_versions = len(db.list_canvases(team_id))
    vp_results = db.list_vp_results(team_id)
    latest_align = vp_results[0]["alignment"] if vp_results else 0.0

    approved_pivots = sum(1 for p in db.list_pivots(team_id)
                          if p["status"] in ("Approved", "Conditional"))
    reflections = len(db.list_reflections(team_id))
    ai_logs = db.list_ai_logs(team_id)
    ai_verified = sum(1 for l in ai_logs if l["status"] in ("Verified", "Modified"))
    ai_rate = (ai_verified / len(ai_logs)) if ai_logs else None

    bmc_data = (bmc or {}).get("data", {}) if bmc else {}
    has_revenue = bool(str(bmc_data.get("revenue_streams", "")).strip())
    has_cost = bool(str(bmc_data.get("cost_structure", "")).strip())
    pricing_exps = sum(1 for e in db.list_experiments(team_id)
                       if "price" in (e["card_type"] or "").lower()
                       or "preorder" in (e["card_type"] or "").lower())

    raw = {}
    raw["Customer Insight"] = (50 * (cp_filled / 3) + min(20, max(0, cp_versions - 1) * 10)
                               + min(30, ev["behavioral"] * 5))
    raw["Value Proposition Fit"] = 50 * (vpc_filled / 6) + 50 * float(latest_align or 0)
    raw["Evidence Strength"] = ev["avg_strength"] * 8 + min(20, ev["behavioral"] * 4)
    raw["Business-Model Coherence"] = 100 * (bmc_filled / 9)
    raw["Experiment Efficiency"] = eff["resolved"] * 20 + (eff["experiments"] - eff["resolved"]) * 5
    raw["Financial Viability"] = ((25 if has_revenue else 0) + (25 if has_cost else 0)
                                  + min(30, risk["supported"] * 10) + (20 if pricing_exps else 0))
    raw["Adaptability"] = all_versions * 8 + approved_pivots * 20
    raw["Responsible Innovation"] = 50 + (50 * ai_rate if ai_rate is not None else 0)
    raw["Team Execution"] = reflections * 15 + min(40, (eff["experiments"] + ev["count"]) * 4)
    # Founder/team skills give a bonus to their mapped dimension.
    for dim, bonus in skill_bonus(team_id).items():
        if dim in raw:
            raw[dim] += bonus
    raw["Investor Confidence"] = (0.4 * raw["Evidence Strength"]
                                  + 0.3 * raw["Business-Model Coherence"]
                                  + 0.3 * raw["Value Proposition Fit"]
                                  - len(risk["exposed"]) * 5)
    return raw


# Each dashboard dimension only becomes fair to score once the tool it depends on
# has been introduced. This keeps Auto-Director recommendations aligned with the
# options actually available in a given round.
_DIMENSION_TOOL = {
    "Customer Insight": "Canvases",
    "Value Proposition Fit": "VP Auction",
    "Evidence Strength": "Evidence Ledger",
    "Experiment Efficiency": "Experiment Marketplace",
    "Financial Viability": "Experiment Marketplace",
    "Adaptability": "Pivot Petition",
    "Responsible Innovation": None,   # AI Assist is a base tool (always available)
    "Team Execution": None,           # Decision Journal is a base tool
}


def dimension_available(dim, round_no):
    """Is a dashboard dimension 'in play' (its tool introduced) by this round?"""
    if dim == "Business-Model Coherence":
        return round_no >= canvas_unlock_round("bmc")
    if dim == "Investor Confidence":
        return round_no >= canvas_unlock_round("bmc")   # needs a business model to judge
    page = _DIMENSION_TOOL.get(dim)
    return page is None or round_no >= page_unlock_round(page)


def available_dimensions(round_no):
    """Dimensions whose underlying tool has been introduced by this round."""
    return [d for d in content.DIMENSION_NAMES if dimension_available(d, round_no)]


def auto_scores(team_id, weights=None, round_no=None, only_available=True):
    """Predict 0–100 dashboard scores from a team's submitted work.

    Each dimension's raw heuristic score is multiplied by the Director's tunable
    weight (default 1.0) and clamped to 0–100. When `round_no` is given and
    `only_available` is True, only dimensions whose tools are introduced by that
    round are returned — so recommendations match what teams could actually do.
    """
    raw = _raw_scores(team_id)
    weights = weights or get_score_weights()
    scores = {dim: _clamp100(raw[dim] * weights.get(dim, 1.0))
              for dim in content.DIMENSION_NAMES}
    if round_no is not None and only_available:
        avail = set(available_dimensions(round_no))
        scores = {d: v for d, v in scores.items() if d in avail}
    return scores


def valuation_from_scores(team_id, scores):
    """Predicted valuation from a set of (possibly not-yet-saved) scores.

    Applies the same evidence-coverage discount as compute_valuation, so a
    predicted valuation can't run ahead of the evidence the team has gathered."""
    team = db.get_team(team_id)
    if not team:
        return None
    potential = (team["market_potential"]
                 * _index_from_score(scores.get("Evidence Strength"))
                 * _index_from_score(scores.get("Business-Model Coherence"))
                 * _index_from_score(scores.get("Team Execution")))
    value = potential * evidence_coverage(team_id) - (team["unresolved_risk"] or 0)
    return round(value, 0)


def apply_scores(team_id, round_no, scores):
    for dim, val in scores.items():
        db.set_score(team_id, round_no, dim, val)


# Map the risk type of a team's biggest exposed assumption to an event category.
_RISK_TO_EVENT = {
    "Desirability": "Customer",
    "Feasibility": "Operational",
    "Viability": "Financial",
    "Adaptability": "Competitive",
}


def suggest_event(team_id, round_no=None):
    """Suggest a market event aimed at the team's biggest exposed risk.

    Returns {category, text, exposes, reason}. Falls back to the round's topic
    when the team has no exposed assumption yet.
    """
    import random
    risk = assumption_risk_report(team_id)
    category = None
    reason = ""
    if risk["exposed"]:
        top = risk["exposed"][0]
        category = _RISK_TO_EVENT.get(top["risk_type"])
        reason = f"Targets untested {top['risk_type']} assumption: “{top['text']}”."
    if not category:
        # Fall back to the current round's canvas focus / topic flavor.
        rnd = round_no or db.current_round()
        focuses = canvas_focus_for_round(rnd)
        if "bmc" in focuses:
            category = "Competitive"
        elif focuses:
            category = "Customer"
        else:
            category = random.choice(list(content.MARKET_EVENTS.keys()))
        reason = "No high-risk untested assumption yet — chosen from this round's focus."
    text, exposes = random.choice(content.MARKET_EVENTS[category])
    return {"category": category, "text": text, "exposes": exposes, "reason": reason}


def recommend_pivot(pivot):
    """Recommend a committee decision for a pivot petition from its completeness."""
    has_evidence = bool((pivot.get("challenge_evid") or "").strip())
    has_change = bool((pivot.get("proposed_change") or "").strip())
    has_new = bool((pivot.get("new_assumptions") or "").strip())
    has_needed = bool((pivot.get("evidence_needed") or "").strip())
    if not has_change or not (pivot.get("original_assum") or "").strip():
        return "Rejected", "Missing the original assumption or a concrete proposed change."
    if not has_evidence:
        return "NeedsEvidence", "No challenging evidence cited — ask the team to test first."
    if has_evidence and has_new and has_needed:
        cost = pivot.get("change_cost") or 0
        if cost and cost > 1000:
            return "Conditional", "Evidence-based, but costly — approve subject to milestones."
        return "Approved", "Evidence cited and new assumptions/tests defined — disciplined pivot."
    return "Conditional", "Partly justified — approve conditionally and request the missing pieces."


# ---- Per-round feedback "email" ------------------------------------------- #
def generate_feedback(team_id, round_no=None, scores=None):
    """Compose a per-round feedback email for a team from its predicted performance.

    Returns {subject, body}. Highlights strengths and weak spots, evidence quality,
    exposed risks, the market event they face, and a concrete next step.
    """
    team = db.get_team(team_id)
    rnd = round_no or db.current_round()
    scores = scores or auto_scores(team_id, round_no=rnd)  # only dimensions in play
    ev = evidence_summary(team_id)
    risk = assumption_risk_report(team_id)

    ranked = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    strengths = [d for d, s in ranked[:3] if s >= 55]
    weak = [d for d, s in ranked[::-1][:3] if s <= 55]
    pred_val = valuation_from_scores(team_id, scores)

    _pi, _pn, _pe, _pt = content.narrative_phase(rnd, total_rounds())
    lines = [f"Hi {team['name']},", "",
             f"{_pe} Chapter {_pi + 1} — {_pn}. {_pt}", "",
             f"Here's your Round {rnd} venture review.", ""]
    lines.append(f"💬 {investor_line(team_id)}")
    lines.append("")
    lines.append(f"Predicted venture valuation: ${pred_val:,.0f}")
    lines.append("")
    if strengths:
        lines.append("What's working:")
        for d in strengths:
            lines.append(f"  • {d} ({scores[d]}/100)")
        lines.append("")
    if weak:
        lines.append("Where to focus next:")
        for d in weak:
            lines.append(f"  • {d} ({scores[d]}/100)")
        lines.append("")

    lines.append(f"Evidence: {ev['count']} logged, {ev['behavioral']} behavioral "
                 f"(avg strength {ev['avg_strength']}/10). "
                 + ("Strong behavioral base — keep it up."
                    if ev['behavioral'] >= 3 else
                    "Push for more behavioral evidence (trials, LOIs, preorders) over opinions."))
    lines.append("")

    if risk["exposed"]:
        lines.append("⚠️ High-risk assumptions you haven't tested yet:")
        for a in risk["exposed"][:3]:
            lines.append(f"  • {a['text']} ({a['risk_type']})")
        top = risk["exposed"][0]
        lines.append("")
        lines.append(f"Recommended next step: design the cheapest experiment that could "
                     f"disprove “{top['text']}”, set a success/failure threshold in advance, "
                     f"and run it before you invest further.")
    else:
        lines.append("Recommended next step: convert your latest learning into your next "
                     "canvas version and line up the next assumption to test.")

    # Reference the newest market event they were dealt, if any.
    events = [e for e in db.list_events(team_id) if e["round"] == rnd]
    if events:
        e = events[0]
        lines.append("")
        lines.append(f"Market watch ({e['category']}): {e['text']} "
                     f"This pressures the assumption: {e['exposes']}")

    hints = round_hints(rnd)
    if hints:
        lines.append("")
        lines.append(f"**To do well in Round {rnd}:**")
        for h in hints:
            lines.append(f"  • {h}")

    lines.append("")
    lines.append(content.INVESTOR["sign"])
    return {"subject": f"Round {rnd} venture review — {team['name']}",
            "body": "\n".join(lines)}


def send_feedback(team_id, round_no=None, scores=None):
    fb = generate_feedback(team_id, round_no, scores)
    db.add_message(team_id, fb["subject"], fb["body"], round_no or db.current_round())
    return fb


# ---- Round hints & the welcome email --------------------------------------- #
def round_hints(rnd):
    """Subtle 'how to do well' hints for the topics scheduled in a round."""
    out = []
    for tp in topics_for_round(rnd):
        h = content.round_hint(tp["key"])
        if h and h not in out:
            out.append(h)
    return out


def founder_tailored_hints(team_id):
    """A few hints tailored to the team's founder card — strengths to lean on and
    gaps to watch."""
    card = db.get_founder_card(team_id)
    levels = content.card_skill_levels(card.get("name", ""))
    lines = []
    if levels:
        strong = max(levels, key=levels.get)
        weak = min(levels, key=levels.get)
        if levels[strong] >= 3:
            s = content.FOUNDER_SKILL_BY_KEY[strong]
            lines.append(f"You're strong in **{s['name']}** — lean on it; it lifts your "
                         f"{s['dimension']}.")
        if levels[weak] <= 1:
            w = content.FOUNDER_SKILL_BY_KEY[weak]
            lines.append(f"You're thin on **{w['name']}** — plan to train it, or hire a "
                         f"{content.SPECIALIST_ROLES.get(weak,'specialist')} when a round leans on it.")
    if card.get("networks"):
        lines.append(f"Your network into {card['networks']} is a head start — begin customer "
                     "discovery with people you can already reach.")
    if card.get("budget"):
        lines.append(f"You can afford to lose ${card['budget']} — spend it on the cheapest tests "
                     "that produce the strongest evidence, not on polish.")
    return lines


def generate_welcome(team_id):
    """A warm Round-1 welcome email with founder- and territory-tailored, subtle hints."""
    team = db.get_team(team_id)
    territory = team["opportunity"] or "your opportunity territory"
    lines = [f"Welcome to Venture Foundry, {team['name']}! 🏭", "",
             "You've been accepted into the accelerator as founders — not with a finished "
             "product, but with an opportunity and a team. Here's the one rule that shapes "
             "everything: you earn resources by producing **credible evidence** that your "
             "business model could work, not by having a good idea or a polished pitch.", ""]
    lines.append(f"**Your opportunity:** {territory}.")
    guide = content.territory_guide(territory)
    if guide:
        lines.append("")
        lines.append("Getting started here — a map, not the answer (go let real customers redraw it):")
        lines.append(guide)
    lines.append("")
    lines.append("**Your founding team.** Read your founder card closely — it's your real "
                 "advantage and your real constraint:")
    for h in founder_tailored_hints(team_id):
        lines.append(f"  • {h}")
    lines.append("")
    hints = round_hints(db.current_round()) or round_hints(1)
    lines.append("**This round, to get off to a strong start:**")
    for h in hints:
        lines.append(f"  • {h}")
    lines.append("")
    lines.append("Set a sustainable time allocation on **Founder & Team**, review your card and "
                 "territory on **Founder & Opportunity**, and check your **Round Briefing** for "
                 "this round's tasks. Good luck — build the evidence.")
    lines.append("")
    lines.append("— The Venture Foundry Director")
    return {"subject": f"Welcome to Venture Foundry — {team['name']}",
            "body": "\n".join(lines)}


def send_welcome(team_id):
    w = generate_welcome(team_id)
    db.add_message(team_id, w["subject"], w["body"], 1)
    return w


# ---- Automation settings & batch run -------------------------------------- #
def auto_flag(key, default=True):
    val = db.get_setting(key)
    if val is None:
        return default
    return str(val) == "1"


def set_auto_flag(key, value):
    db.set_setting(key, "1" if value else "0")


def run_autopilot(round_no=None, teams=None):
    """Apply enabled automation for a round. Returns a per-team summary.

    Idempotent-ish: scores overwrite for the round; an event is only issued to a
    team that has none for that round; pending pivots are decided by recommendation.
    """
    rnd = round_no or db.current_round()
    teams = teams if teams is not None else db.list_teams()
    events_open = rnd >= page_unlock_round("Market Events")
    pivots_open = rnd >= page_unlock_round("Pivot Petition")
    summary = []
    for t in teams:
        entry = {"team": t["name"], "team_id": t["id"], "scored": False,
                 "event": None, "pivots": 0}
        if auto_flag("auto_scoring_on"):
            sc = auto_scores(t["id"], round_no=rnd)   # only dimensions in play this round
            apply_scores(t["id"], rnd, sc)
            entry["scored"] = True
        if auto_flag("auto_events_on") and events_open:
            existing = [e for e in db.list_events(t["id"], include_broadcast=False)
                        if e["round"] == rnd]
            if not existing:
                ev = suggest_event(t["id"], rnd)
                db.add_event(t["id"], rnd, ev["category"], ev["text"], ev["exposes"])
                entry["event"] = ev["category"]
        if auto_flag("auto_pivots_on") and pivots_open:
            for p in db.list_pivots(t["id"]):
                if p["status"] == "Submitted":
                    dec, note = recommend_pivot(p)
                    db.decide_pivot(p["id"], dec, "[auto] " + note)
                    if dec in ("Approved", "Conditional") and (p["change_cost"] or 0):
                        db.adjust_resources(t["id"], money=-p["change_cost"], kind="pivot",
                                            description="Pivot change cost (auto)",
                                            allow_negative=True)
                    entry["pivots"] += 1
        if auto_flag("auto_feedback_on"):
            send_feedback(t["id"], rnd)
            entry["feedback"] = True
        summary.append(entry)
    return summary
